"""Test _generate_ltc_excel directly to find the 500 error."""
import sys, io, json
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / 'Rel Website' / 'backend'))
sys.path.insert(0, str(Path(__file__).parent))

req = {
    'request_number': 'REL202600001',
    'original_rr_number': 'REL202600001',
    'product_hierarchy': 'PKG/LEAD',
    'classification': 'Standard',
    'originator': 'TestUser',
    'pdl': 'PDL1',
    'body_size_x': 8.0,
    'body_size_y': 8.0,
    'plant': 'K1',
    'package_thickness': 1.2,
    'device_name': 'TEST-DEVICE',
    'lot_no': 'LOT001',
    'ball_pitch': 0.5,
    'customer': 'AMKOR',
    'ball_count': 256,
    'pkg_info': 'BGA',
    'lead_pitch': 0.4,
    'automotive': False,
    'lead_count': 0,
    'total_ss': 10,
    'purpose': 'Qualification Test',
    'steps': [
        {
            'leg': 1, 'step_number': 1, 'step_name': 'HAST 96h',
            'custom_fields': {
                'test_item': 'Electrical', 
                'test_condition': '130C/85%RH/96h',
                'reading_pt': 'T0/T96'
            }
        },
        {
            'leg': 1, 'step_number': 2, 'step_name': 'uSAT',
            'custom_fields': {'test_item': 'Delamination', 'test_condition': '', 'reading_pt': 'T0/T96'}
        }
    ]
}

# Run the core of _generate_ltc_excel inline
import zipfile, re
from openpyxl import load_workbook
from datetime import date

TPL = Path('Rel Website/backend/templates/REL LTC Template.xlsx')
print(f'Template: {TPL} exists={TPL.exists()}')

try:
    wb = load_workbook(str(TPL))
    ws = wb['Sheet1']
    today = date.today()

    def _w(cell_ref, value):
        if value is None or value == '': return
        ws[cell_ref].value = value

    _w('D8',  req.get('request_number', ''))
    _w('F8',  req.get('original_rr_number', '') or req.get('request_number', ''))  # Automated Rel Number
    _w('O8',  req.get('product_hierarchy', ''))
    _w('D9',  req.get('classification', ''))
    _w('O9',  req.get('pdl', ''))
    _w('D10', req.get('originator', ''))
    _w('D11', req.get('plant', ''))
    _w('D12', req.get('device_name', ''))
    _w('D13', req.get('lot_no', ''))
    _w('D14', req.get('customer', ''))
    _w('D15', req.get('pkg_info', ''))
    _w('D16', 'Yes' if req.get('automotive') else 'No')
    ws['D17'].value = today
    _w('O17', req.get('total_ss', ''))
    _w('D18', req.get('purpose', ''))
    print('header writes OK')

    steps = req.get('steps', [])
    steps_sorted = sorted(steps, key=lambda s: (s.get('leg', 1), s.get('step_number', 0)))
    lot_no = req.get('lot_no', '')
    total_ss = req.get('total_ss', '')
    MATRIX_START = 26

    for i, step in enumerate(steps_sorted):
        r = MATRIX_START + i
        cf = step.get('custom_fields', {}) or {}
        _w(f'B{r}', step.get('leg', 1))
        _w(f'C{r}', lot_no)
        _w(f'F{r}', step.get('step_name', ''))
        _w(f'H{r}', cf.get('test_item', ''))
        _w(f'J{r}', cf.get('test_condition', ''))
        _w(f'M{r}', cf.get('reading_pt', '') or cf.get('reading_point', ''))
        _w(f'T{r}', total_ss)
    print('matrix writes OK')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    print('wb.save OK')

    with zipfile.ZipFile(buf) as zf:
        s1 = zf.read('xl/worksheets/sheet1.xml')
        s1 = re.sub(rb'<drawing\b[^>]*/>', b'<drawing r:id="rId2"/>', s1)
        ss = zf.read('xl/sharedStrings.xml') if 'xl/sharedStrings.xml' in zf.namelist() else None
    print('openpyxl extract OK')

    output = io.BytesIO()
    with zipfile.ZipFile(str(TPL)) as tpl_zip, \
         zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as out_zip:
        for name in tpl_zip.namelist():
            if name == 'xl/worksheets/sheet1.xml':
                out_zip.writestr(name, s1)
            elif name == 'xl/sharedStrings.xml' and ss:
                out_zip.writestr(name, ss)
            elif name == 'xl/calcChain.xml':
                pass
            elif name == '[Content_Types].xml':
                ct = tpl_zip.read(name)
                ct = re.sub(rb'<Override\s+PartName="/xl/calcChain\.xml"[^>]*/>', b'', ct)
                out_zip.writestr(name, ct)
            elif name == 'xl/_rels/workbook.xml.rels':
                rels = tpl_zip.read(name)
                rels = re.sub(rb'<Relationship\s[^>]*calcChain[^>]*/>', b'', rels)
                out_zip.writestr(name, rels)
            else:
                out_zip.writestr(name, tpl_zip.read(name))
    print('ZIP merge OK')

    output.seek(0)
    data = output.getvalue()
    Path('test_ltc_output.xlsx').write_bytes(data)
    print(f'SUCCESS - {len(data)} bytes -> test_ltc_output.xlsx')

    # Verify
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        names = z.namelist()
        ct = z.read('[Content_Types].xml').decode()
        rl = z.read('xl/_rels/workbook.xml.rels').decode()
        s1d = z.read('xl/worksheets/sheet1.xml').decode()
        drw = re.search(r'<drawing\b[^>]*/>', s1d)
        print(f'  calcChain in CT:  {"calcChain" in ct}')
        print(f'  calcChain in rels: {"calcChain" in rl}')
        print(f'  calcChain in ZIP:  {"xl/calcChain.xml" in names}')
        print(f'  docMetadata:       {any("docMetadata" in n for n in names)}')
        print(f'  Drawing ref:       {drw.group(0) if drw else "NOT FOUND"}')

except Exception as e:
    import traceback
    traceback.print_exc()
    print(f'\nERROR: {e}')

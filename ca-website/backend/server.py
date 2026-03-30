"""
CA (Construction Analysis) Request Process Flow — Backend Server
Port: 8001  |  Database: ca_database.db  (isolated from REL)
"""

import os, jwt, bcrypt, sqlite3, aiosqlite, logging, io, json
from pathlib import Path
from datetime import datetime, timedelta, timezone
from typing import Optional, List
from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException, Depends, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel, EmailStr
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

ROOT_DIR = Path(__file__).parent

# ── Config ─────────────────────────────────────────────────────────────────
SECRET_KEY   = os.getenv("CA_SECRET_KEY", "ca-secret-key-change-in-prod")
ALGORITHM    = "HS256"
TOKEN_EXPIRY = 24  # hours
DB_PATH      = os.path.join(os.path.dirname(__file__), "ca_database.db")
# Port: Support both CA_PORT and PORT environment variables (CA_PORT takes precedence)
PORT         = int(os.getenv("CA_PORT") or os.getenv("PORT", 8001))
# Network: Listen on all interfaces for LAN sharing (can be overridden with HOST env var)
HOST         = os.getenv("HOST", "0.0.0.0")

# Path to the Rel (REL Request) database for user sync
REL_DB_PATH  = os.getenv(
    "REL_DB_PATH",
    str(ROOT_DIR.parent.parent / "Rel Website" / "backend" / "rel_database.db")
)

# Role mapping: Rel roles → CA roles
REL_TO_CA_ROLE: dict = {
    "Admin":                "Admin",
    "Reliability Engineer": "REL Engineer",
    "Failure Analysis":     "REL Engineer",
    "Technician":           "Technician",
    "Planner":              "Planner",
    "Requestor":            "Analyst",
}

def map_rel_role(rel_role: str) -> str:
    """Return the closest CA role for a given Rel role."""
    return REL_TO_CA_ROLE.get(rel_role, "Analyst")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("ca_server")

CA_STEPS_DEFAULT = [
    "External Visual Inspection",
    "X-ray Inspection",
    "SAT Inspection",
    "Chemical Decapsulation",
    "Optical Inspection",
    "Laser & Plasma Decapsulation",
    "Wire Pull Test",
    "Ball Shear Test",
    "Stitch Pull Test",
    "SEM Inspection",
    "Manual Cross-Section (SEM)",
    "Ion Mill",
]

ROLES = ["Admin", "REL Engineer", "Analyst", "Technician", "Planner"]

# Default check items for each CA analysis step
CHECKLIST_TEMPLATE = [
    {"step": "External Visual Inspection", "items": [
        {"name": "top including marking", "requirements": "Check for mold/leadframe discoloration, mold flash, pealing or any anomaly to be heads up to Rel Engr.", "qty": "", "remarks": ""},
        {"name": "bottom", "requirements": "", "qty": "", "remarks": ""},
        {"name": "top leads", "requirements": "", "qty": "", "remarks": ""},
        {"name": "bottom leads", "requirements": "", "qty": "", "remarks": ""},
        {"name": "side leads", "requirements": "Scribe on Epad side for traceability", "qty": "", "remarks": ""},
    ]},
    {"step": "X-ray Inspection", "items": [
        {"name": "mold", "requirements": "Check for insufficient die attach coverage or any wire anomaly to be heads up to Rel Engr. Unit mark Traceability", "qty": "", "remarks": ""},
        {"name": "die attach", "requirements": "", "qty": "", "remarks": ""},
        {"name": "wires", "requirements": "", "qty": "", "remarks": ""},
    ]},
    {"step": "SAT Inspection", "items": [
        {"name": "thru-scan", "requirements": "Any unit with abnormality/void more than 300um measurement must NOT BE DESTRUCTED and to be heads up to Rel Engr.", "qty": "", "remarks": ""},
        {"name": "die top", "requirements": "", "qty": "", "remarks": ""},
        {"name": "paddle top", "requirements": "Unit mark as traceability; color map 24 and map 1 for 3-scan; 4 units maximum per Scan with nail polish", "qty": "", "remarks": ""},
        {"name": "paddle bottom", "requirements": "", "qty": "-", "remarks": ""},
        {"name": "leadframe", "requirements": "", "qty": "", "remarks": ""},
        {"name": "others: Void Measurement", "requirements": "", "qty": "", "remarks": "MAX VOID ONLY IF ANY"},
    ]},
    {"step": "Chemical Decapsulation", "items": [
        {"name": "", "requirements": "", "qty": "", "remarks": ""},
    ]},
    {"step": "Optical Inspection", "items": [
        {"name": "unit profile", "requirements": "Unit Traceability and magnification per image", "qty": "", "remarks": ""},
        {"name": "die surface", "requirements": "", "qty": "", "remarks": ""},
        {"name": "die metallization", "requirements": "", "qty": "", "remarks": ""},
        {"name": "Ball Shear Breakmode", "requirements": "", "qty": "", "remarks": ""},
        {"name": "others: Invalidated reading", "requirements": "", "qty": "", "remarks": ""},
    ]},
    {"step": "Laser & Plasma Decapsulation", "items": [
        {"name": "", "requirements": "", "qty": "", "remarks": ""},
    ]},
    {"step": "Wire Pull Test", "items": [
        {"name": "all bonds", "requirements": "Spec Limit: 6 gF", "qty": "", "remarks": ""},
        {"name": "others:", "requirements": "", "qty": "", "remarks": ""},
    ]},
    {"step": "Ball Shear Test", "items": [
        {"name": "all bonds", "requirements": "Spec Limit: 17 gF", "qty": "", "remarks": ""},
        {"name": "others:", "requirements": "", "qty": "", "remarks": ""},
    ]},
    {"step": "Stitch Pull Test", "items": [
        {"name": "all bonds", "requirements": "Spec Limit: 4 gF", "qty": "-", "remarks": ""},
        {"name": "others:", "requirements": "", "qty": "", "remarks": ""},
    ]},
    {"step": "SEM Inspection", "items": [
        {"name": "unit profile", "requirements": "Unit & Pad Traceability", "qty": "", "remarks": ""},
        {"name": "loop formation (Corner Photo)", "requirements": "SEM Image per Abnormality", "qty": "", "remarks": ""},
        {"name": "ball/bond formation/placement", "requirements": "Highest and lowest reading Breakmode SEM Image", "qty": "", "remarks": ""},
        {"name": "stitchbond formation", "requirements": "", "qty": "", "remarks": ""},
        {"name": "wpt break mode", "requirements": "", "qty": "", "remarks": ""},
        {"name": "bst break mode", "requirements": "", "qty": "", "remarks": ""},
        {"name": "spt break mode", "requirements": "", "qty": "", "remarks": ""},
        {"name": "others: Defects", "requirements": "", "qty": "", "remarks": "stitch shear (RSSB) die side"},
    ]},
    {"step": "Manual Cross-Section (SEM)", "items": [
        {"name": "Actual Cross Section", "requirements": "obvious & suspected hairline Gap", "qty": "", "remarks": ""},
        {"name": "Ball Diameter", "requirements": "", "qty": "", "remarks": ""},
        {"name": "Ball Height", "requirements": "", "qty": "", "remarks": ""},
        {"name": "Ball Bond Gap Measurement", "requirements": "", "qty": "", "remarks": ""},
        {"name": "Wedge Formation Inspection", "requirements": "", "qty": "", "remarks": ""},
        {"name": "Wedge Thickness", "requirements": "", "qty": "", "remarks": ""},
        {"name": "others:", "requirements": "", "qty": "", "remarks": ""},
    ]},
    {"step": "Ion Mill", "items": [
        {"name": "Decoration", "requirements": "", "qty": "", "remarks": ""},
    ]},
]

# ── Database Init ──────────────────────────────────────────────────────────
async def init_db():
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                email       TEXT UNIQUE NOT NULL,
                username    TEXT NOT NULL,
                password    TEXT NOT NULL,
                role        TEXT DEFAULT 'Analyst',
                is_approved INTEGER DEFAULT 1,
                is_active   INTEGER DEFAULT 1,
                blocked     INTEGER DEFAULT 0,
                user_status TEXT DEFAULT 'approved',
                last_seen   TEXT,
                avatar      TEXT,
                created_at  TEXT DEFAULT (datetime('now'))
            )
        """)
        # Add columns to the users table if the DB already exists without them
        for col, definition in [
            ("blocked",     "INTEGER DEFAULT 0"),
            ("user_status", "TEXT DEFAULT 'approved'"),
            ("last_seen",   "TEXT"),
            ("avatar",      "TEXT"),
        ]:
            try:
                await db.execute(f"ALTER TABLE users ADD COLUMN {col} {definition}")
            except Exception:
                pass  # Column already exists
        await db.execute("""
            CREATE TABLE IF NOT EXISTS login_logs (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id    INTEGER NOT NULL,
                username   TEXT NOT NULL,
                email      TEXT NOT NULL,
                role       TEXT NOT NULL,
                login_at   TEXT DEFAULT (datetime('now'))
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS role_permissions (
                role       TEXT NOT NULL,
                permission TEXT NOT NULL,
                enabled    INTEGER DEFAULT 1,
                PRIMARY KEY (role, permission)
            )
        """)
        # Seed default CA role permissions if table is empty
        cur = await db.execute("SELECT COUNT(*) FROM role_permissions")
        if (await cur.fetchone())[0] == 0:
            default_perms = [
                ("REL Engineer", "create_request", 1),
                ("REL Engineer", "edit_request", 1),
                ("REL Engineer", "delete_request", 0),
                ("REL Engineer", "update_steps", 1),
                ("REL Engineer", "manage_steps", 0),
                ("REL Engineer", "manage_users", 0),
                ("REL Engineer", "manage_settings", 0),
                ("Analyst", "create_request", 1),
                ("Analyst", "edit_request", 0),
                ("Analyst", "delete_request", 0),
                ("Analyst", "update_steps", 1),
                ("Analyst", "manage_steps", 0),
                ("Analyst", "manage_users", 0),
                ("Analyst", "manage_settings", 0),
                ("Technician", "create_request", 0),
                ("Technician", "edit_request", 0),
                ("Technician", "delete_request", 0),
                ("Technician", "update_steps", 1),
                ("Technician", "manage_steps", 0),
                ("Technician", "manage_users", 0),
                ("Technician", "manage_settings", 0),
                ("Planner", "create_request", 1),
                ("Planner", "edit_request", 1),
                ("Planner", "delete_request", 0),
                ("Planner", "update_steps", 1),
                ("Planner", "manage_steps", 1),
                ("Planner", "manage_users", 0),
                ("Planner", "manage_settings", 0),
            ]
            await db.executemany(
                "INSERT OR IGNORE INTO role_permissions (role, permission, enabled) VALUES (?,?,?)",
                default_perms
            )
        await db.commit()
        # Add employee tracking columns to login_logs (migration-safe)
        for _col, _def in [
            ("employee_id",   "TEXT DEFAULT ''"),
            ("employee_name", "TEXT DEFAULT ''"),
        ]:
            try:
                await db.execute(f"ALTER TABLE login_logs ADD COLUMN {_col} {_def}")
            except Exception:
                pass  # Column already exists
        await db.execute("""
            CREATE TABLE IF NOT EXISTS technician_sessions (
                employee_id       TEXT PRIMARY KEY,
                employee_name     TEXT NOT NULL DEFAULT '',
                employee_position TEXT NOT NULL DEFAULT '',
                last_active       TEXT NOT NULL,
                login_at          TEXT NOT NULL
            )
        """)
        await db.commit()
        await db.execute("""
            CREATE TABLE IF NOT EXISTS ca_requests (
                id                 INTEGER PRIMARY KEY AUTOINCREMENT,
                ca_number          TEXT UNIQUE NOT NULL,
                title              TEXT NOT NULL,
                sample_description TEXT,
                lot_number         TEXT,
                device             TEXT,
                department         TEXT,
                submitter_id       INTEGER,
                submitter_name     TEXT,
                submitter_email    TEXT,
                analyst_id         INTEGER,
                status             TEXT DEFAULT 'pending',
                priority           TEXT DEFAULT 'Normal',
                note               TEXT,
                current_step       TEXT,
                discontinue_reason TEXT,
                created_at         TEXT DEFAULT (datetime('now')),
                approved_at        TEXT,
                completed_at       TEXT,
                discontinued_at    TEXT,
                -- General Information fields
                classification     TEXT,
                originator         TEXT,
                plant              TEXT,
                device_name        TEXT,
                lot_no             TEXT,
                customer           TEXT,
                pkg_info           TEXT,
                automotive         TEXT,
                reference_project  TEXT,
                product_hierarchy  TEXT,
                pdl                TEXT,
                body_size_x        TEXT,
                body_size_y        TEXT,
                package_thickness  TEXT,
                ball_pitch         TEXT,
                ball_count         TEXT,
                lead_pitch         TEXT,
                lead_count         TEXT,
                total_ss           TEXT,
                purpose            TEXT,
                -- Material Information fields
                bcb_material                   TEXT,
                bump_height                    TEXT,
                bump_material                  TEXT,
                bump_pitch                     TEXT,
                bump_size                      TEXT,
                bumping_house                  TEXT,
                chip_attach_flux_cleaning_method TEXT,
                chip_attach_flux               TEXT,
                die_attach_material            TEXT,
                die_coat_after_wb              TEXT,
                die_pad_config                 TEXT,
                die_pad_metal                  TEXT,
                die_pad_pitch                  TEXT,
                die_passivation                TEXT,
                die_size                       TEXT,
                die_thick                      TEXT,
                down_bond                      TEXT,
                emc_encap_material             TEXT,
                heat_dissipation_matl          TEXT,
                lf_ag_option                   TEXT,
                lf_etch_stamp                  TEXT,
                lf_inner_lead_pitch            TEXT,
                lf_sub_material                TEXT,
                lf_sub_pad_size                TEXT,
                lf_sub_supplier                TEXT,
                lf_sub_thickness               TEXT,
                lid_attach_epoxy               TEXT,
                line_width                     TEXT,
                mfg_site                       TEXT,
                masking_material               TEXT,
                others1                        TEXT,
                others2                        TEXT,
                others3                        TEXT,
                others4                        TEXT,
                others5                        TEXT,
                passive_component              TEXT,
                pcb_finish                     TEXT,
                plating_option                 TEXT,
                rel_site                       TEXT,
                solder_ball_attach_paste       TEXT,
                solder_ball_material           TEXT,
                solder_ball_size               TEXT,
                solder_mask_material           TEXT,
                solder_paste_material          TEXT,
                sub_layer                      TEXT,
                sub_pad_design                 TEXT,
                sub_pad_opening_size           TEXT,
                sub_surface_treatment          TEXT,
                ubm_material                   TEXT,
                ubm_opening_size               TEXT,
                underfill_material             TEXT,
                wafer_type                     TEXT,
                wire_length_max                TEXT,
                wire_material                  TEXT,
                wire_size                      TEXT,
                wire_supplier                  TEXT,
                wire_type                      TEXT,
                FOREIGN KEY (submitter_id) REFERENCES users(id),
                FOREIGN KEY (analyst_id)   REFERENCES users(id)
            )
        """)
        # Migrate existing ca_requests table — add new columns if absent
        _new_ca_cols = [
            ("classification", "TEXT"), ("originator", "TEXT"), ("plant", "TEXT"),
            ("device_name", "TEXT"), ("lot_no", "TEXT"), ("customer", "TEXT"),
            ("pkg_info", "TEXT"), ("automotive", "TEXT"), ("reference_project", "TEXT"),
            ("product_hierarchy", "TEXT"), ("pdl", "TEXT"),
            ("body_size_x", "TEXT"), ("body_size_y", "TEXT"),
            ("package_thickness", "TEXT"), ("ball_pitch", "TEXT"), ("ball_count", "TEXT"),
            ("lead_pitch", "TEXT"), ("lead_count", "TEXT"), ("total_ss", "TEXT"),
            ("purpose", "TEXT"),
            ("bcb_material", "TEXT"), ("bump_height", "TEXT"), ("bump_material", "TEXT"),
            ("bump_pitch", "TEXT"), ("bump_size", "TEXT"), ("bumping_house", "TEXT"),
            ("chip_attach_flux_cleaning_method", "TEXT"), ("chip_attach_flux", "TEXT"),
            ("die_attach_material", "TEXT"), ("die_coat_after_wb", "TEXT"),
            ("die_pad_config", "TEXT"), ("die_pad_metal", "TEXT"), ("die_pad_pitch", "TEXT"),
            ("die_passivation", "TEXT"), ("die_size", "TEXT"), ("die_thick", "TEXT"),
            ("down_bond", "TEXT"), ("emc_encap_material", "TEXT"),
            ("heat_dissipation_matl", "TEXT"), ("lf_ag_option", "TEXT"),
            ("lf_etch_stamp", "TEXT"), ("lf_inner_lead_pitch", "TEXT"),
            ("lf_sub_material", "TEXT"), ("lf_sub_pad_size", "TEXT"),
            ("lf_sub_supplier", "TEXT"), ("lf_sub_thickness", "TEXT"),
            ("lid_attach_epoxy", "TEXT"), ("line_width", "TEXT"),
            ("mfg_site", "TEXT"), ("masking_material", "TEXT"),
            ("others1", "TEXT"), ("others2", "TEXT"), ("others3", "TEXT"),
            ("others4", "TEXT"), ("others5", "TEXT"), ("passive_component", "TEXT"),
            ("pcb_finish", "TEXT"), ("plating_option", "TEXT"), ("rel_site", "TEXT"),
            ("solder_ball_attach_paste", "TEXT"), ("solder_ball_material", "TEXT"),
            ("solder_ball_size", "TEXT"), ("solder_mask_material", "TEXT"),
            ("solder_paste_material", "TEXT"), ("sub_layer", "TEXT"),
            ("sub_pad_design", "TEXT"), ("sub_pad_opening_size", "TEXT"),
            ("sub_surface_treatment", "TEXT"), ("ubm_material", "TEXT"),
            ("ubm_opening_size", "TEXT"), ("underfill_material", "TEXT"),
            ("wafer_type", "TEXT"), ("wire_length_max", "TEXT"),
            ("wire_material", "TEXT"), ("wire_size", "TEXT"),
            ("wire_supplier", "TEXT"), ("wire_type", "TEXT"),
        ]
        for col, coltype in _new_ca_cols:
            try:
                await db.execute(f"ALTER TABLE ca_requests ADD COLUMN {col} {coltype}")
            except Exception:
                pass  # column already exists
        # Add due_date to ca_requests if not present
        try:
            await db.execute("ALTER TABLE ca_requests ADD COLUMN due_date TEXT")
        except Exception:
            pass
        # Add retention_details column to ca_requests if not present
        try:
            await db.execute("ALTER TABLE ca_requests ADD COLUMN retention_details TEXT")
        except Exception:
            pass
        await db.execute("""
            CREATE TABLE IF NOT EXISTS ca_steps (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id   INTEGER NOT NULL,
                step_number  INTEGER NOT NULL,
                step_name    TEXT NOT NULL,
                status       TEXT DEFAULT 'not_started',
                remarks      TEXT,
                started_at   TEXT,
                completed_at TEXT,
                assigned_to  INTEGER,
                FOREIGN KEY (request_id)  REFERENCES ca_requests(id) ON DELETE CASCADE,
                FOREIGN KEY (assigned_to) REFERENCES users(id)
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS ca_schedule (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id   INTEGER NOT NULL,
                step_name    TEXT,
                analyst_id   INTEGER,
                scheduled_date TEXT,
                due_date       TEXT,
                notes          TEXT,
                status         TEXT DEFAULT 'scheduled',
                created_at     TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (request_id) REFERENCES ca_requests(id) ON DELETE CASCADE,
                FOREIGN KEY (analyst_id) REFERENCES users(id)
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key   TEXT PRIMARY KEY,
                value TEXT
            )
        """)
        await db.execute("INSERT OR IGNORE INTO settings (key,value) VALUES ('maintenance_mode','false')")
        await db.execute("INSERT OR IGNORE INTO settings (key,value) VALUES ('site_name','CA Request Process Flow')")
        await db.execute("""
            CREATE TABLE IF NOT EXISTS ca_checklist_items (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id   INTEGER NOT NULL,
                leg_name     TEXT DEFAULT '',
                leg_title    TEXT DEFAULT '',
                step_name    TEXT NOT NULL,
                sort_order   INTEGER NOT NULL DEFAULT 0,
                item_name    TEXT DEFAULT '',
                requirements TEXT DEFAULT '',
                time_in      TEXT DEFAULT '',
                time_out     TEXT DEFAULT '',
                technician   TEXT DEFAULT '',
                qty          TEXT DEFAULT '',
                remarks      TEXT DEFAULT '',
                FOREIGN KEY (request_id) REFERENCES ca_requests(id) ON DELETE CASCADE
            )
        """)
        # Migrate: add leg columns to existing DB if missing
        for _col, _def in [("leg_name", "TEXT DEFAULT ''"), ("leg_title", "TEXT DEFAULT ''")]:
            try:
                await db.execute(f"ALTER TABLE ca_checklist_items ADD COLUMN {_col} {_def}")
            except Exception:
                pass  # column already exists
        await db.commit()

    # Seed admin user if empty
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM users")
    if cur.fetchone()[0] == 0:
        pw = bcrypt.hashpw(b"admin123", bcrypt.gensalt()).decode()
        cur.execute(
            "INSERT INTO users (email,username,password,role,is_approved) VALUES (?,?,?,?,1)",
            ("admin@amkor.com", "Admin", pw, "Admin"),
        )
        conn.commit()
        logger.info("Seeded admin: admin@amkor.com / admin123")
    conn.close()

# ── Lifespan & App ─────────────────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    logger.info(f"CA Server running on port {PORT}")
    yield

app = FastAPI(title="CA Request API", version="1.0.0", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True,
                   allow_methods=["*"], allow_headers=["*"])

security = HTTPBearer(auto_error=False)

# ── Auth helpers ───────────────────────────────────────────────────────────
def make_token(uid: int, role: str) -> str:
    exp = datetime.now(timezone.utc) + timedelta(hours=TOKEN_EXPIRY)
    return jwt.encode({"sub": str(uid), "role": role, "exp": exp}, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(creds: HTTPAuthorizationCredentials = Depends(security)):
    if not creds:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(creds.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        # Guest token — no DB lookup needed
        if payload.get("is_guest"):
            emp_id   = payload.get("employee_id", "")
            emp_name = payload.get("employee_name", "")
            display  = emp_name or "Technician"
            return {"id": 0, "email": "guest@ca.local", "username": display,
                    "role": "Technician", "is_active": 1, "is_approved": 1, "is_guest": True,
                    "employee_id": emp_id, "employee_name": emp_name}
        uid = int(payload["sub"])
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute("SELECT * FROM users WHERE id=? AND is_active=1", (uid,))
        u = await cur.fetchone()
    if not u:
        raise HTTPException(status_code=401, detail="User not found")
    return dict(u)

def require_role(*roles):
    async def checker(u=Depends(get_current_user)):
        if u["role"] not in roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return u
    return checker

# ── Pydantic models ────────────────────────────────────────────────────────
class LoginData(BaseModel):
    email: EmailStr
    password: str

class RegisterData(BaseModel):
    email: EmailStr
    username: str
    password: str
    role: str = "Analyst"

class RequestCreate(BaseModel):
    title: str
    sample_description: Optional[str] = None
    lot_number: Optional[str] = None
    device: Optional[str] = None
    department: Optional[str] = None
    submitter_name: Optional[str] = None
    priority: str = "Normal"
    note: Optional[str] = None
    # General Information
    classification: Optional[str] = None
    originator: Optional[str] = None
    plant: Optional[str] = None
    device_name: Optional[str] = None
    lot_no: Optional[str] = None
    customer: Optional[str] = None
    pkg_info: Optional[str] = None
    automotive: Optional[str] = None
    reference_project: Optional[str] = None
    product_hierarchy: Optional[str] = None
    pdl: Optional[str] = None
    body_size_x: Optional[str] = None
    body_size_y: Optional[str] = None
    package_thickness: Optional[str] = None
    ball_pitch: Optional[str] = None
    ball_count: Optional[str] = None
    lead_pitch: Optional[str] = None
    lead_count: Optional[str] = None
    total_ss: Optional[str] = None
    purpose: Optional[str] = None
    # Material Information
    bcb_material: Optional[str] = None
    bump_height: Optional[str] = None
    bump_material: Optional[str] = None
    bump_pitch: Optional[str] = None
    bump_size: Optional[str] = None
    bumping_house: Optional[str] = None
    chip_attach_flux_cleaning_method: Optional[str] = None
    chip_attach_flux: Optional[str] = None
    die_attach_material: Optional[str] = None
    die_coat_after_wb: Optional[str] = None
    die_pad_config: Optional[str] = None
    die_pad_metal: Optional[str] = None
    die_pad_pitch: Optional[str] = None
    die_passivation: Optional[str] = None
    die_size: Optional[str] = None
    die_thick: Optional[str] = None
    down_bond: Optional[str] = None
    emc_encap_material: Optional[str] = None
    heat_dissipation_matl: Optional[str] = None
    lf_ag_option: Optional[str] = None
    lf_etch_stamp: Optional[str] = None
    lf_inner_lead_pitch: Optional[str] = None
    lf_sub_material: Optional[str] = None
    lf_sub_pad_size: Optional[str] = None
    lf_sub_supplier: Optional[str] = None
    lf_sub_thickness: Optional[str] = None
    lid_attach_epoxy: Optional[str] = None
    line_width: Optional[str] = None
    mfg_site: Optional[str] = None
    masking_material: Optional[str] = None
    others1: Optional[str] = None
    others2: Optional[str] = None
    others3: Optional[str] = None
    others4: Optional[str] = None
    others5: Optional[str] = None
    passive_component: Optional[str] = None
    pcb_finish: Optional[str] = None
    plating_option: Optional[str] = None
    rel_site: Optional[str] = None
    solder_ball_attach_paste: Optional[str] = None
    solder_ball_material: Optional[str] = None
    solder_ball_size: Optional[str] = None
    solder_mask_material: Optional[str] = None
    solder_paste_material: Optional[str] = None
    sub_layer: Optional[str] = None
    sub_pad_design: Optional[str] = None
    sub_pad_opening_size: Optional[str] = None
    sub_surface_treatment: Optional[str] = None
    ubm_material: Optional[str] = None
    ubm_opening_size: Optional[str] = None
    underfill_material: Optional[str] = None
    wafer_type: Optional[str] = None
    wire_length_max: Optional[str] = None
    wire_material: Optional[str] = None
    wire_size: Optional[str] = None
    wire_supplier: Optional[str] = None
    wire_type: Optional[str] = None

class StepUpdate(BaseModel):
    status: Optional[str] = None
    remarks: Optional[str] = None
    assigned_to: Optional[int] = None

class ChecklistItemUpdate(BaseModel):
    time_in: Optional[str] = None
    time_out: Optional[str] = None
    technician: Optional[str] = None
    qty: Optional[str] = None
    remarks: Optional[str] = None

class ChecklistBulkItem(BaseModel):
    leg_name: str = ''
    leg_title: str = ''
    step_name: str = ''
    item_name: str = ''
    requirements: str = ''
    qty: str = ''
    remarks: str = ''

class UserUpdate(BaseModel):
    role: Optional[str] = None
    is_active: Optional[bool] = None
    is_approved: Optional[bool] = None

class ScheduleCreate(BaseModel):
    request_id: int
    step_name: Optional[str] = None
    analyst_id: Optional[int] = None
    scheduled_date: Optional[str] = None
    due_date: Optional[str] = None
    notes: Optional[str] = None

class ScheduleUpdate(BaseModel):
    analyst_id: Optional[int] = None
    scheduled_date: Optional[str] = None
    due_date: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None

class DiscontinueBody(BaseModel):
    reason: Optional[str] = None

class ApproveBody(BaseModel):
    due_date: str

# ── Utilities ─────────────────────────────────────────────────────────────
async def next_ca_number() -> str:
    year = datetime.now().year
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "SELECT ca_number FROM ca_requests WHERE ca_number LIKE ? ORDER BY id DESC LIMIT 1",
            (f"CA-{year}-%",)
        )
        row = await cur.fetchone()
    seq = (int(row[0].split("-")[-1]) + 1) if row else 1
    return f"CA-{year}-{seq:04d}"

# ── Public ─────────────────────────────────────────────────────────────────
@app.get("/api/health")
async def health():
    return {"status": "ok", "service": "CA Request API", "port": PORT}

@app.get("/api/public/stats")
async def public_stats():
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT COUNT(*) FROM users WHERE is_approved=1 AND is_active=1")
        count = (await cur.fetchone())[0]
    return {"approved_users": count}

# ── Auth ───────────────────────────────────────────────────────────────────
@app.post("/api/auth/login")
async def login(body: LoginData):
    # 1. Try the CA database first
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute("SELECT * FROM users WHERE email=? AND is_active=1", (body.email,))
        u = await cur.fetchone()
    if u and bcrypt.checkpw(body.password.encode(), u["password"].encode()):
        if not u["is_approved"]:
            raise HTTPException(status_code=403, detail="Account not yet approved")
        token = make_token(u["id"], u["role"])
        now = datetime.now(timezone.utc).isoformat()
        async with aiosqlite.connect(DB_PATH) as db2:
            await db2.execute(
                "INSERT INTO login_logs (user_id, username, email, role, login_at) VALUES (?,?,?,?,?)",
                (u["id"], u["username"], u["email"], u["role"], now)
            )
            await db2.execute("UPDATE users SET last_seen=? WHERE id=?", (now, u["id"]))
            await db2.commit()
        return {"token": token, "user": {"id": u["id"], "email": u["email"],
                "username": u["username"], "role": u["role"]}}

    # 2. Fallback: check the Rel database and auto-import user if credentials match
    rel_db = os.path.normpath(REL_DB_PATH)
    if os.path.exists(rel_db):
        try:
            conn = sqlite3.connect(rel_db)
            conn.row_factory = sqlite3.Row
            rel_row = conn.execute(
                "SELECT email, username, password, role, approved FROM users WHERE email=? AND approved=1",
                (body.email,)
            ).fetchone()
            conn.close()

            if rel_row and bcrypt.checkpw(body.password.encode(), rel_row["password"].encode()):
                ca_role = map_rel_role(rel_row["role"])
                async with aiosqlite.connect(DB_PATH) as db:
                    # Re-check: user might exist but inactive
                    ec = await db.execute("SELECT id FROM users WHERE email=?", (body.email,))
                    existing = await ec.fetchone()
                    if existing:
                        await db.execute(
                            "UPDATE users SET username=?, password=?, role=?, is_active=1, is_approved=1 WHERE email=?",
                            (rel_row["username"], rel_row["password"], ca_role, body.email)
                        )
                        uid = existing[0]
                    else:
                        ic = await db.execute(
                            "INSERT INTO users (email,username,password,role,is_approved,is_active) VALUES (?,?,?,?,1,1)",
                            (rel_row["email"], rel_row["username"], rel_row["password"], ca_role)
                        )
                        uid = ic.lastrowid
                    await db.commit()
                now = datetime.now(timezone.utc).isoformat()
                async with aiosqlite.connect(DB_PATH) as db3:
                    await db3.execute(
                        "INSERT INTO login_logs (user_id, username, email, role, login_at) VALUES (?,?,?,?,?)",
                        (uid, rel_row["username"], rel_row["email"], ca_role, now)
                    )
                    await db3.execute("UPDATE users SET last_seen=? WHERE id=?", (now, uid))
                    await db3.commit()
                token = make_token(uid, ca_role)
                return {"token": token, "user": {"id": uid, "email": rel_row["email"],
                        "username": rel_row["username"], "role": ca_role}}
        except Exception as exc:
            logger.warning(f"Rel DB fallback login error: {exc}")

    raise HTTPException(status_code=401, detail="Invalid email or password")

@app.post("/api/auth/register")
async def register(body: RegisterData):
    if body.role not in ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    pw = bcrypt.hashpw(body.password.encode(), bcrypt.gensalt()).decode()
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            cur = await db.execute(
                "INSERT INTO users (email,username,password,role,is_approved) VALUES (?,?,?,?,1)",
                (body.email, body.username, pw, body.role)
            )
            uid = cur.lastrowid
            await db.commit()
    except Exception:
        raise HTTPException(status_code=400, detail="Email already registered")
    token = make_token(uid, body.role)
    return {"token": token, "user": {"id": uid, "email": body.email,
            "username": body.username, "role": body.role}}

@app.get("/api/auth/me")
async def me(u=Depends(get_current_user)):
    return {k: u[k] for k in ("id", "email", "username", "role")}

@app.post("/api/auth/heartbeat")
async def heartbeat(u=Depends(get_current_user)):
    now = datetime.now(timezone.utc).isoformat()
    if u.get("is_guest"):
        emp_id = u.get("employee_id", "")
        if emp_id:
            async with aiosqlite.connect(DB_PATH) as db:
                await db.execute(
                    "UPDATE technician_sessions SET last_active=? WHERE employee_id=?",
                    (now, emp_id)
                )
                await db.commit()
        return {"ok": True}
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE users SET last_seen=? WHERE id=?", (now, u["id"]))
        await db.commit()
    return {"ok": True}

@app.post("/api/verify-tech-code")
async def verify_tech_code(data: dict):
    """Public endpoint — verify the 6-digit Technician passcode."""
    code = data.get("code", "")
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT value FROM settings WHERE key='tech_auth_code'")
        row = await cur.fetchone()
    stored = row[0] if row and row[0] else "735522"
    return {"valid": code == stored}

class GuestTokenData(BaseModel):
    employee_id: Optional[str] = None
    employee_name: Optional[str] = None
    employee_position: Optional[str] = None

@app.post("/api/auth/guest-token")
async def guest_token(body: GuestTokenData = None):
    """Issue a short-lived JWT for a guest Technician — no credentials required."""
    emp_id  = (body.employee_id or "").strip() if body else ""
    emp_name = (body.employee_name or "").strip() if body else ""
    emp_pos  = (body.employee_position or "").strip() if body else ""
    display  = emp_name or "Technician"
    exp = datetime.now(timezone.utc) + timedelta(hours=TOKEN_EXPIRY)
    token = jwt.encode(
        {"sub": "guest", "is_guest": True, "exp": exp,
         "employee_id": emp_id, "employee_name": emp_name},
        SECRET_KEY, algorithm=ALGORITHM
    )
    now = datetime.now(timezone.utc).isoformat()
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT INTO login_logs (user_id, username, email, role, login_at, employee_id, employee_name) VALUES (?,?,?,?,?,?,?)",
            (0, display, "guest@ca.local", "Technician", now, emp_id, emp_name)
        )
        if emp_id:
            await db.execute("""
                INSERT INTO technician_sessions (employee_id, employee_name, employee_position, last_active, login_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(employee_id) DO UPDATE SET
                    employee_name=excluded.employee_name,
                    employee_position=excluded.employee_position,
                    last_active=excluded.last_active,
                    login_at=excluded.login_at
            """, (emp_id, emp_name, emp_pos, now, now))
        await db.commit()
    return {
        "token": token,
        "user": {
            "id": "guest",
            "email": "guest@ca.local",
            "username": display,
            "role": "Technician",
            "is_guest": True,
            "employee_id": emp_id,
            "employee_name": emp_name,
        }
    }

# ── Role Permissions ───────────────────────────────────────────────────────
ALL_PERMISSIONS = [
    "create_request", "edit_request", "delete_request",
    "update_steps", "manage_steps", "manage_users", "manage_settings",
]
CONFIGURABLE_ROLES_CA = ["REL Engineer", "Analyst", "Technician", "Planner"]

@app.get("/api/role-permissions")
async def get_role_permissions(_=Depends(get_current_user)):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute("SELECT role, permission, enabled FROM role_permissions")
        rows = await cur.fetchall()
    perms: dict = {r: {} for r in CONFIGURABLE_ROLES_CA}
    for row in rows:
        if row["role"] in perms:
            perms[row["role"]][row["permission"]] = bool(row["enabled"])
    # Ensure every permission key exists (default False)
    for role in CONFIGURABLE_ROLES_CA:
        for perm in ALL_PERMISSIONS:
            perms[role].setdefault(perm, False)
    return {"permissions": perms, "all_permissions": ALL_PERMISSIONS}

class RolePermissionsUpdate(BaseModel):
    permissions: dict

@app.put("/api/role-permissions")
async def update_role_permissions(body: RolePermissionsUpdate, _=Depends(require_role("Admin"))):
    async with aiosqlite.connect(DB_PATH) as db:
        for role, perms in body.permissions.items():
            if role not in CONFIGURABLE_ROLES_CA:
                continue
            for perm, enabled in perms.items():
                await db.execute(
                    "INSERT OR REPLACE INTO role_permissions (role, permission, enabled) VALUES (?,?,?)",
                    (role, perm, 1 if enabled else 0)
                )
        await db.commit()
    return {"message": "Permissions updated"}

@app.get("/api/my-permissions")
async def my_permissions(u=Depends(get_current_user)):
    if u["role"] == "Admin":
        return {p: True for p in ALL_PERMISSIONS}
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute(
            "SELECT permission, enabled FROM role_permissions WHERE role=?", (u["role"],)
        )
        rows = await cur.fetchall()
    result = {p: False for p in ALL_PERMISSIONS}
    for row in rows:
        result[row["permission"]] = bool(row["enabled"])
    return result

# ── Users ──────────────────────────────────────────────────────────────────
def _map_user_row(r: dict) -> dict:
    """Normalise a CA user row to match what the REL-style Users page expects."""
    blocked = bool(r.get("blocked", 0))
    approved = bool(r.get("is_approved", 1))
    status = r.get("user_status") or ("lock" if blocked else "approved" if approved else "pending")
    return {
        "id":          r["id"],
        "email":       r["email"],
        "username":    r["username"],
        "role":        r["role"],
        "approved":    approved,
        "blocked":     blocked,
        "user_status": status,
        "is_active":   bool(r.get("is_active", 1)),
        "last_seen":   r.get("last_seen"),
        "avatar":      r.get("avatar"),
        "created_at":  r.get("created_at"),
    }

@app.get("/api/users")
async def list_users(_=Depends(get_current_user)):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute(
            "SELECT id,email,username,role,is_approved,is_active,blocked,user_status,last_seen,avatar,created_at "
            "FROM users ORDER BY created_at DESC"
        )
        rows = await cur.fetchall()
    return [_map_user_row(dict(r)) for r in rows]

@app.patch("/api/users/{uid}")
async def update_user(uid: int, body: UserUpdate, _=Depends(require_role("Admin"))):
    fields, values = [], []
    if body.role is not None:
        if body.role not in ROLES:
            raise HTTPException(status_code=400, detail="Invalid role")
        fields.append("role=?"); values.append(body.role)
    if body.is_active is not None:
        fields.append("is_active=?"); values.append(1 if body.is_active else 0)
    if body.is_approved is not None:
        fields.append("is_approved=?"); values.append(1 if body.is_approved else 0)
    if not fields:
        return {"message": "Nothing to update"}
    values.append(uid)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(f"UPDATE users SET {', '.join(fields)} WHERE id=?", values)
        await db.commit()
    return {"message": "User updated"}

@app.patch("/api/users/{uid}/approve")
async def approve_user(uid: int, _=Depends(require_role("Admin"))):
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "UPDATE users SET is_approved=1, user_status='approved', blocked=0 WHERE id=?", (uid,)
        )
        await db.commit()
    return {"message": "User approved"}

@app.patch("/api/users/{uid}/reject")
async def reject_user(uid: int, _=Depends(require_role("Admin"))):
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "UPDATE users SET is_approved=0, user_status='declined' WHERE id=?", (uid,)
        )
        await db.commit()
    return {"message": "User rejected"}

class RoleBody(BaseModel):
    role: str

@app.patch("/api/users/{uid}/role")
async def update_user_role(uid: int, body: RoleBody, _=Depends(require_role("Admin"))):
    if body.role not in ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE users SET role=? WHERE id=?", (body.role, uid))
        await db.commit()
    return {"message": "Role updated"}

class StatusBody(BaseModel):
    status: str

@app.patch("/api/users/{uid}/status")
async def update_user_status(uid: int, body: StatusBody, _=Depends(require_role("Admin"))):
    valid = {"approved", "hold", "lock", "declined", "pending"}
    if body.status not in valid:
        raise HTTPException(status_code=400, detail="Invalid status")
    blocked = 1 if body.status == "lock" else 0
    approved = 1 if body.status == "approved" else 0
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "UPDATE users SET user_status=?, blocked=?, is_approved=? WHERE id=?",
            (body.status, blocked, approved, uid)
        )
        await db.commit()
    return {"message": "Status updated"}

class UsernameBody(BaseModel):
    username: str

@app.patch("/api/users/{uid}/username")
async def update_user_username(uid: int, body: UsernameBody, _=Depends(require_role("Admin"))):
    if not body.username.strip():
        raise HTTPException(status_code=400, detail="Username cannot be empty")
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE users SET username=? WHERE id=?", (body.username.strip(), uid))
        await db.commit()
    return {"message": "Username updated"}

@app.patch("/api/users/{uid}/block")
async def toggle_block_user(uid: int, current=Depends(require_role("Admin"))):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute("SELECT blocked, is_approved FROM users WHERE id=?", (uid,))
        row = await cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="User not found")
        new_blocked = 0 if row["blocked"] else 1
        new_status = "lock" if new_blocked else ("approved" if row["is_approved"] else "pending")
        await db.execute(
            "UPDATE users SET blocked=?, user_status=? WHERE id=?",
            (new_blocked, new_status, uid)
        )
        await db.commit()
    return {"message": "User block status toggled"}

@app.delete("/api/users/{uid}")
async def delete_user(uid: int, current=Depends(require_role("Admin"))):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute("SELECT user_status, is_approved, blocked FROM users WHERE id=?", (uid,))
        row = await cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="User not found")
        status = row["user_status"] or ("lock" if row["blocked"] else "approved" if row["is_approved"] else "pending")
        if status == "approved":
            raise HTTPException(status_code=400, detail="Cannot delete an approved user")
        await db.execute("DELETE FROM users WHERE id=?", (uid,))
        await db.commit()
    return {"message": "User deleted"}

@app.get("/api/login-logs")
async def get_login_logs(_=Depends(require_role("Admin"))):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute(
            "SELECT id,user_id,username,email,role,login_at,"
            "COALESCE(employee_id,'') AS employee_id,COALESCE(employee_name,'') AS employee_name "
            "FROM login_logs ORDER BY login_at DESC LIMIT 500"
        )
        rows = await cur.fetchall()
    return [dict(r) for r in rows]

@app.get("/api/active-technicians")
async def get_active_technicians(_=Depends(get_current_user)):
    """Return technician sessions active in the last 5 minutes."""
    five_min_ago = (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute(
            "SELECT employee_id, employee_name, employee_position, last_active, login_at "
            "FROM technician_sessions WHERE last_active >= ? ORDER BY last_active DESC",
            (five_min_ago,)
        )
        rows = await cur.fetchall()
    return [dict(r) for r in rows]

@app.post("/api/admin/sync-rel-users")
async def sync_rel_users(_=Depends(require_role("Admin"))):
    """Copy all approved Rel users into the CA database.
    Passwords are bcrypt-compatible so users can log in with their existing credentials.
    Rel roles are mapped to the closest CA role.
    """
    rel_db = os.path.normpath(REL_DB_PATH)
    if not os.path.exists(rel_db):
        raise HTTPException(status_code=404, detail=f"Rel database not found at: {rel_db}")
    try:
        conn = sqlite3.connect(rel_db)
        conn.row_factory = sqlite3.Row
        rel_users = conn.execute(
            "SELECT email, username, password, role FROM users WHERE approved=1"
        ).fetchall()
        conn.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read Rel database: {e}")

    created, updated = 0, 0
    async with aiosqlite.connect(DB_PATH) as db:
        for row in rel_users:
            ca_role = map_rel_role(row["role"])
            ec = await db.execute("SELECT id FROM users WHERE email=?", (row["email"],))
            existing = await ec.fetchone()
            if existing:
                await db.execute(
                    "UPDATE users SET username=?, password=?, role=?, is_active=1, is_approved=1 WHERE email=?",
                    (row["username"], row["password"], ca_role, row["email"])
                )
                updated += 1
            else:
                await db.execute(
                    "INSERT INTO users (email,username,password,role,is_approved,is_active) VALUES (?,?,?,?,1,1)",
                    (row["email"], row["username"], row["password"], ca_role)
                )
                created += 1
        await db.commit()

    return {
        "message": f"Sync complete. {created} user(s) created, {updated} user(s) updated.",
        "created": created,
        "updated": updated,
        "total": created + updated,
    }

# ── CA Requests ────────────────────────────────────────────────────────────
async def _build_requests(rows, db_path):
    result = []
    for r in rows:
        d = dict(r)
        async with aiosqlite.connect(db_path) as db2:
            db2.row_factory = aiosqlite.Row
            sc = await db2.execute(
                "SELECT * FROM ca_steps WHERE request_id=? ORDER BY step_number", (d["id"],)
            )
            d["steps"] = [dict(s) for s in await sc.fetchall()]
        result.append(d)
    return result

@app.get("/api/requests")
async def list_requests(_=Depends(get_current_user)):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute("""
            SELECT r.*, u.username AS analyst_name
            FROM ca_requests r
            LEFT JOIN users u ON r.analyst_id=u.id
            ORDER BY r.created_at DESC
        """)
        rows = await cur.fetchall()
    return await _build_requests(rows, DB_PATH)

@app.get("/api/requests/{req_id}")
async def get_request(req_id: int, _=Depends(get_current_user)):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute("""
            SELECT r.*, u.username AS analyst_name
            FROM ca_requests r
            LEFT JOIN users u ON r.analyst_id=u.id
            WHERE r.id=?
        """, (req_id,))
        row = await cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Request not found")
    d = dict(row)
    async with aiosqlite.connect(DB_PATH) as db2:
        db2.row_factory = aiosqlite.Row
        sc = await db2.execute(
            "SELECT * FROM ca_steps WHERE request_id=? ORDER BY step_number", (req_id,)
        )
        d["steps"] = [dict(s) for s in await sc.fetchall()]
    return d

@app.get("/api/requests/{req_id}/checklist")
async def get_checklist(req_id: int, _=Depends(get_current_user)):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute("SELECT COUNT(*) as cnt FROM ca_checklist_items WHERE request_id=?", (req_id,))
        row = await cur.fetchone()
        if row["cnt"] == 0:
            sort = 0
            for tpl in CHECKLIST_TEMPLATE:
                for item in tpl["items"]:
                    sort += 1
                    await db.execute(
                        "INSERT INTO ca_checklist_items (request_id,step_name,sort_order,item_name,requirements,qty,remarks) VALUES (?,?,?,?,?,?,?)",
                        (req_id, tpl["step"], sort, item["name"], item["requirements"], item["qty"], item["remarks"])
                    )
            await db.commit()
        cur = await db.execute(
            "SELECT * FROM ca_checklist_items WHERE request_id=? ORDER BY sort_order", (req_id,)
        )
        rows = await cur.fetchall()
    return [dict(r) for r in rows]

@app.post("/api/requests/{req_id}/checklist/bulk")
async def import_checklist_bulk(req_id: int, items: List[ChecklistBulkItem], _=Depends(get_current_user)):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT id FROM ca_requests WHERE id=?", (req_id,))
        if not await cur.fetchone():
            raise HTTPException(status_code=404, detail="Request not found")
        await db.execute("DELETE FROM ca_checklist_items WHERE request_id=?", (req_id,))
        for i, item in enumerate(items, start=1):
            await db.execute(
                "INSERT INTO ca_checklist_items (request_id,leg_name,leg_title,step_name,sort_order,item_name,requirements,qty,remarks) VALUES (?,?,?,?,?,?,?,?,?)",
                (req_id, item.leg_name, item.leg_title, item.step_name, i, item.item_name, item.requirements, item.qty, item.remarks)
            )
        await db.commit()
    return {"imported": len(items)}

@app.patch("/api/checklist/{item_id}")
async def update_checklist_item(item_id: int, body: ChecklistItemUpdate, _=Depends(get_current_user)):
    fields, values = [], []
    for field, val in [
        ("time_in", body.time_in), ("time_out", body.time_out),
        ("technician", body.technician), ("qty", body.qty), ("remarks", body.remarks),
    ]:
        if val is not None:
            fields.append(f"{field}=?")
            values.append(val)
    if not fields:
        return {"message": "Nothing to update"}
    values.append(item_id)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(f"UPDATE ca_checklist_items SET {', '.join(fields)} WHERE id=?", values)
        await db.commit()
    return {"message": "Updated"}

@app.post("/api/requests")
async def create_request(body: RequestCreate, u=Depends(get_current_user)):
    ca_num = await next_ca_number()
    now = datetime.now(timezone.utc).isoformat()
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("""
            INSERT INTO ca_requests
              (ca_number,title,sample_description,lot_number,device,department,
               submitter_id,submitter_name,submitter_email,priority,note,current_step,
               classification,originator,plant,device_name,lot_no,customer,pkg_info,
               automotive,reference_project,product_hierarchy,pdl,
               body_size_x,body_size_y,package_thickness,ball_pitch,ball_count,
               lead_pitch,lead_count,total_ss,purpose,
               bcb_material,bump_height,bump_material,bump_pitch,bump_size,bumping_house,
               chip_attach_flux_cleaning_method,chip_attach_flux,die_attach_material,
               die_coat_after_wb,die_pad_config,die_pad_metal,die_pad_pitch,die_passivation,
               die_size,die_thick,down_bond,emc_encap_material,heat_dissipation_matl,
               lf_ag_option,lf_etch_stamp,lf_inner_lead_pitch,lf_sub_material,lf_sub_pad_size,
               lf_sub_supplier,lf_sub_thickness,lid_attach_epoxy,line_width,
               mfg_site,masking_material,others1,others2,others3,others4,others5,
               passive_component,pcb_finish,plating_option,rel_site,
               solder_ball_attach_paste,solder_ball_material,solder_ball_size,
               solder_mask_material,solder_paste_material,sub_layer,sub_pad_design,
               sub_pad_opening_size,sub_surface_treatment,ubm_material,ubm_opening_size,
               underfill_material,wafer_type,wire_length_max,wire_material,wire_size,
               wire_supplier,wire_type)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,
                    ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,
                    ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            ca_num, body.title, body.sample_description, body.lot_number,
            body.device, body.department, u["id"],
            body.submitter_name or u["username"], u["email"],
            body.priority, body.note, CA_STEPS_DEFAULT[0],
            body.classification, body.originator, body.plant,
            body.device_name, body.lot_no, body.customer, body.pkg_info,
            body.automotive, body.reference_project, body.product_hierarchy, body.pdl,
            body.body_size_x, body.body_size_y, body.package_thickness,
            body.ball_pitch, body.ball_count, body.lead_pitch, body.lead_count,
            body.total_ss, body.purpose,
            body.bcb_material, body.bump_height, body.bump_material,
            body.bump_pitch, body.bump_size, body.bumping_house,
            body.chip_attach_flux_cleaning_method, body.chip_attach_flux,
            body.die_attach_material, body.die_coat_after_wb,
            body.die_pad_config, body.die_pad_metal, body.die_pad_pitch,
            body.die_passivation, body.die_size, body.die_thick,
            body.down_bond, body.emc_encap_material, body.heat_dissipation_matl,
            body.lf_ag_option, body.lf_etch_stamp, body.lf_inner_lead_pitch,
            body.lf_sub_material, body.lf_sub_pad_size, body.lf_sub_supplier,
            body.lf_sub_thickness, body.lid_attach_epoxy, body.line_width,
            body.mfg_site, body.masking_material,
            body.others1, body.others2, body.others3, body.others4, body.others5,
            body.passive_component, body.pcb_finish, body.plating_option, body.rel_site,
            body.solder_ball_attach_paste, body.solder_ball_material, body.solder_ball_size,
            body.solder_mask_material, body.solder_paste_material,
            body.sub_layer, body.sub_pad_design, body.sub_pad_opening_size,
            body.sub_surface_treatment, body.ubm_material, body.ubm_opening_size,
            body.underfill_material, body.wafer_type, body.wire_length_max,
            body.wire_material, body.wire_size, body.wire_supplier, body.wire_type,
        ))
        req_id = cur.lastrowid
        for i, sn in enumerate(CA_STEPS_DEFAULT, 1):
            await db.execute(
                "INSERT INTO ca_steps (request_id,step_number,step_name,status) VALUES (?,?,?,'not_started')",
                (req_id, i, sn)
            )
        sort = 0
        for tpl in CHECKLIST_TEMPLATE:
            for item in tpl["items"]:
                sort += 1
                await db.execute(
                    "INSERT INTO ca_checklist_items (request_id,step_name,sort_order,item_name,requirements,qty,remarks) VALUES (?,?,?,?,?,?,?)",
                    (req_id, tpl["step"], sort, item["name"], item["requirements"], item["qty"], item["remarks"])
                )
        await db.commit()
    return {"id": req_id, "ca_number": ca_num}

@app.post("/api/requests/{req_id}/approve")
@app.patch("/api/requests/{req_id}/approve")
async def approve_request(req_id: int, body: ApproveBody, _=Depends(require_role("Admin", "REL Engineer"))):
    if not body.due_date or not body.due_date.strip():
        raise HTTPException(status_code=422, detail="A deadline (due date) is required before approving.")
    now = datetime.now(timezone.utc).isoformat()
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "UPDATE ca_requests SET status='in_progress', approved_at=?, due_date=? WHERE id=?",
            (now, body.due_date.strip(), req_id)
        )
        await db.execute("""
            UPDATE ca_steps SET status='in_progress', started_at=?
            WHERE request_id=? AND step_number=(SELECT MIN(step_number) FROM ca_steps WHERE request_id=?)
        """, (now, req_id, req_id))
        await db.commit()
    return {"message": "Approved"}

@app.post("/api/requests/{req_id}/discontinue")
@app.patch("/api/requests/{req_id}/discontinue")
async def discontinue_request(req_id: int, body: Optional[DiscontinueBody] = None, _=Depends(get_current_user)):
    now = datetime.now(timezone.utc).isoformat()
    reason = (body.reason if body else None) or ""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "UPDATE ca_requests SET status='discontinued', discontinued_at=?, discontinue_reason=? WHERE id=?",
            (now, reason, req_id)
        )
        await db.commit()
    return {"message": "Discontinued"}

@app.patch("/api/requests/{req_id}")
async def update_request(req_id: int, body: dict, _=Depends(get_current_user)):
    """Update editable fields on a CA request (e.g. retention_details)."""
    allowed = {'retention_details', 'note'}
    fields, values = [], []
    for key, val in body.items():
        if key in allowed:
            fields.append(f"{key}=?")
            values.append(val)
    if not fields:
        return {"message": "Nothing to update"}
    values.append(req_id)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(f"UPDATE ca_requests SET {', '.join(fields)} WHERE id=?", values)
        await db.commit()
    return {"message": "Updated"}

@app.delete("/api/requests/{req_id}")
async def delete_request(req_id: int, _=Depends(require_role("Admin"))):
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("DELETE FROM ca_requests WHERE id=?", (req_id,))
        await db.commit()
    return {"message": "Deleted"}

# ── Steps ──────────────────────────────────────────────────────────────────
@app.patch("/api/steps/{step_id}")
async def update_step(step_id: int, body: StepUpdate, _=Depends(get_current_user)):
    fields, values = [], []
    now = datetime.now(timezone.utc).isoformat()
    if body.status is not None:
        fields.append("status=?"); values.append(body.status)
        if body.status == "in_progress":
            fields.append("started_at=?"); values.append(now)
        elif body.status in ("completed", "failed"):
            fields.append("completed_at=?"); values.append(now)
    if body.remarks is not None:
        fields.append("remarks=?"); values.append(body.remarks)
    if body.assigned_to is not None:
        fields.append("assigned_to=?"); values.append(body.assigned_to)
    if not fields:
        return {"message": "Nothing to update"}
    values.append(step_id)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(f"UPDATE ca_steps SET {', '.join(fields)} WHERE id=?", values)
        if body.status == "completed":
            db.row_factory = aiosqlite.Row
            cur = await db.execute("SELECT request_id, step_number FROM ca_steps WHERE id=?", (step_id,))
            row = await cur.fetchone()
            if row:
                nxt = await db.execute(
                    "SELECT id FROM ca_steps WHERE request_id=? AND step_number=? AND status='not_started'",
                    (row["request_id"], row["step_number"] + 1)
                )
                nxt_row = await nxt.fetchone()
                if nxt_row:
                    await db.execute("UPDATE ca_steps SET status='in_queue' WHERE id=?", (nxt_row["id"],))
                    # Update current_step on request
                    step_cur = await db.execute("SELECT step_name FROM ca_steps WHERE id=?", (nxt_row["id"],))
                    step_row = await step_cur.fetchone()
                    if step_row:
                        await db.execute(
                            "UPDATE ca_requests SET current_step=? WHERE id=?",
                            (step_row["step_name"], row["request_id"])
                        )
        await db.commit()
    return {"message": "Step updated"}

# ── Schedule ───────────────────────────────────────────────────────────────
@app.get("/api/schedule")
async def list_schedule(_=Depends(get_current_user)):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute("""
            SELECT s.*, r.ca_number, r.title AS request_title, u.username AS analyst_name
            FROM ca_schedule s
            LEFT JOIN ca_requests r ON s.request_id=r.id
            LEFT JOIN users u ON s.analyst_id=u.id
            ORDER BY s.scheduled_date ASC
        """)
        rows = await cur.fetchall()
    return [dict(r) for r in rows]

@app.post("/api/schedule")
async def create_schedule(body: ScheduleCreate, _=Depends(require_role("Admin", "REL Engineer", "Planner"))):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("""
            INSERT INTO ca_schedule (request_id,step_name,analyst_id,scheduled_date,due_date,notes)
            VALUES (?,?,?,?,?,?)
        """, (body.request_id, body.step_name, body.analyst_id,
              body.scheduled_date, body.due_date, body.notes))
        sid = cur.lastrowid
        await db.commit()
    return {"id": sid}

@app.patch("/api/schedule/{sid}")
async def update_schedule(sid: int, body: ScheduleUpdate, _=Depends(require_role("Admin", "REL Engineer", "Planner"))):
    fields, values = [], []
    if body.analyst_id is not None:
        fields.append("analyst_id=?"); values.append(body.analyst_id)
    if body.scheduled_date is not None:
        fields.append("scheduled_date=?"); values.append(body.scheduled_date)
    if body.due_date is not None:
        fields.append("due_date=?"); values.append(body.due_date)
    if body.notes is not None:
        fields.append("notes=?"); values.append(body.notes)
    if body.status is not None:
        fields.append("status=?"); values.append(body.status)
    if not fields:
        return {"message": "Nothing to update"}
    values.append(sid)
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(f"UPDATE ca_schedule SET {', '.join(fields)} WHERE id=?", values)
        await db.commit()
    return {"message": "Schedule updated"}

@app.delete("/api/schedule/{sid}")
async def delete_schedule(sid: int, _=Depends(require_role("Admin", "REL Engineer", "Planner"))):
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("DELETE FROM ca_schedule WHERE id=?", (sid,))
        await db.commit()
    return {"message": "Deleted"}

# ── Dashboard ──────────────────────────────────────────────────────────────
@app.get("/api/dashboard")
async def dashboard(_=Depends(get_current_user)):
    async with aiosqlite.connect(DB_PATH) as db:
        async def count(q, *a):
            cur = await db.execute(q, a)
            return (await cur.fetchone())[0]

        total        = await count("SELECT COUNT(*) FROM ca_requests")
        pending      = await count("SELECT COUNT(*) FROM ca_requests WHERE status='pending'")
        in_progress  = await count("SELECT COUNT(*) FROM ca_requests WHERE status='in_progress'")
        completed    = await count("SELECT COUNT(*) FROM ca_requests WHERE status='completed'")
        discontinued = await count("SELECT COUNT(*) FROM ca_requests WHERE status='discontinued'")

        db.row_factory = aiosqlite.Row
        cur = await db.execute("""
            SELECT s.id AS step_id, s.step_name, s.request_id,
                   r.ca_number, r.title, r.created_at,
                   GROUP_CONCAT(DISTINCT CASE WHEN ci.leg_name != '' THEN ci.leg_name ELSE NULL END) AS leg_names
            FROM ca_steps s
            JOIN ca_requests r ON s.request_id = r.id
            LEFT JOIN ca_checklist_items ci ON ci.request_id = r.id
            WHERE s.status = 'in_progress'
            GROUP BY s.id
            ORDER BY r.created_at ASC
        """)
        active_steps = [
            {
                "step_name":  r["step_name"],
                "request_id": r["request_id"],
                "ca_number":  r["ca_number"],
                "title":      r["title"],
                "leg_names":  sorted(set(l.strip() for l in (r["leg_names"] or "").split(",") if l.strip())),
            }
            for r in await cur.fetchall()
        ]

        cur2 = await db.execute("""
            SELECT id,ca_number,title,sample_description,submitter_name,status,created_at,current_step
            FROM ca_requests ORDER BY created_at DESC LIMIT 8
        """)
        recent = [dict(r) for r in await cur2.fetchall()]

        # Upcoming schedule items (next 7 days)
        cur3 = await db.execute("""
            SELECT s.*, r.ca_number, r.title AS request_title, u.username AS analyst_name
            FROM ca_schedule s
            LEFT JOIN ca_requests r ON s.request_id=r.id
            LEFT JOIN users u ON s.analyst_id=u.id
            WHERE s.status='scheduled'
            ORDER BY s.scheduled_date ASC LIMIT 5
        """)
        upcoming = [dict(r) for r in await cur3.fetchall()]

        # Status distribution (for bar chart)
        _status_labels = {
            'pending': 'Pending', 'in_progress': 'In Progress',
            'completed': 'Completed', 'discontinued': 'Discontinued',
            'approved': 'Approved',
        }
        cur_sd = await db.execute("SELECT status, COUNT(*) AS cnt FROM ca_requests GROUP BY status")
        status_distribution = [
            {'status': _status_labels.get(r['status'], r['status']), 'count': r['cnt']}
            for r in await cur_sd.fetchall()
        ]

        # Delayed requests: due_date is in the past, still active
        cur_del = await db.execute("""
            SELECT id, ca_number, title, device, lot_number, purpose, due_date
            FROM ca_requests
            WHERE due_date IS NOT NULL
              AND due_date < date('now')
              AND status NOT IN ('completed','discontinued')
            ORDER BY due_date ASC LIMIT 10
        """)
        delayed_requests = [dict(r) for r in await cur_del.fetchall()]

        # Upcoming deadline requests: due_date within next 7 days, still active
        cur_upc = await db.execute("""
            SELECT id, ca_number, title, device, lot_number, purpose, due_date
            FROM ca_requests
            WHERE due_date IS NOT NULL
              AND due_date BETWEEN date('now') AND date('now','+7 days')
              AND status NOT IN ('completed','discontinued')
            ORDER BY due_date ASC LIMIT 10
        """)
        upcoming_deadlines = [dict(r) for r in await cur_upc.fetchall()]

        # Completion rate (last 30 days)
        completed_30d = await count(
            "SELECT COUNT(*) FROM ca_requests WHERE completed_at >= datetime('now','-30 days')"
        )
        total_30d = await count(
            "SELECT COUNT(*) FROM ca_requests WHERE created_at >= datetime('now','-30 days')"
        )
        rate_pct = round(completed_30d / total_30d * 100) if total_30d > 0 else 0

    return {
        "stats": {
            "total": total, "pending": pending, "in_progress": in_progress,
            "completed": completed, "discontinued": discontinued,
        },
        "active_steps": active_steps,
        "recent_requests": recent,
        "upcoming_schedule": upcoming,
        "status_distribution": status_distribution,
        "delayed_requests": delayed_requests,
        "upcoming_deadlines": upcoming_deadlines,
        "completion_rate": {
            "percent": rate_pct,
            "completed_30d": completed_30d,
            "total_30d": total_30d,
        },
    }

# ── Record Monitor ─────────────────────────────────────────────────────────
@app.get("/api/records")
async def list_records(_=Depends(get_current_user)):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute("""
            SELECT r.*, u.username AS analyst_name
            FROM ca_requests r
            LEFT JOIN users u ON r.analyst_id=u.id
            WHERE r.status IN ('completed','discontinued')
            ORDER BY r.completed_at DESC, r.discontinued_at DESC
        """)
        rows = await cur.fetchall()
    return await _build_requests(rows, DB_PATH)

# ── Backup / Export ────────────────────────────────────────────────────────
@app.get("/api/backup/export")
async def export_backup(_=Depends(require_role("Admin"))):
    """Export all completed CA requests as a multi-sheet Excel backup, then
    permanently delete them from the database.

    Sheet 1 – Summary: one row per request (sorted by created_at)
    Sheet 2+ – one sheet per request named by ca_number, showing:
        • General info table at top
        • Per-leg checklist items beneath
    """
    if not _OPENPYXL_OK:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row

        # Fetch COMPLETED requests only
        cur = await db.execute("""
            SELECT r.*, u.username AS analyst_name
            FROM ca_requests r
            LEFT JOIN users u ON r.analyst_id = u.id
            WHERE r.status = 'completed'
            ORDER BY r.created_at ASC
        """)
        requests = [dict(r) for r in await cur.fetchall()]

        # Fetch checklist items for each request
        for req in requests:
            c2 = await db.execute(
                "SELECT * FROM ca_checklist_items WHERE request_id=? ORDER BY sort_order",
                (req["id"],)
            )
            req["checklist"] = [dict(r) for r in await c2.fetchall()]
            c3 = await db.execute(
                "SELECT * FROM ca_steps WHERE request_id=? ORDER BY step_number",
                (req["id"],)
            )
            req["steps"] = [dict(r) for r in await c3.fetchall()]

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "CA Requests Summary"

    # ── Styles ──────────────────────────────────────────────────────────
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill   = PatternFill("solid", fgColor="4C1D95")   # violet-900
    sub_fill   = PatternFill("solid", fgColor="EDE9FE")   # violet-100
    sub_font   = Font(bold=True, color="4C1D95", size=9)
    leg_fill   = PatternFill("solid", fgColor="F3F4F6")   # gray-100
    leg_font   = Font(bold=True, color="1E1B4B", size=9)
    thin_side  = Side(style="thin", color="C4B5FD")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_al  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def _hdr(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = center_al; c.border = thin_border
        if width:
            ws.column_dimensions[c.column_letter].width = width
        return c

    def _cell(ws, row, col, val):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = left_al; c.border = thin_border
        return c

    # ── Sheet 1: Summary ────────────────────────────────────────────────
    SUM_COLS = [
        ("CA Number", 16), ("Title", 38), ("Status", 14),
        ("Submitter", 20), ("Analyst", 20), ("Priority", 12),
        ("Created At", 18), ("Completed At", 18),
        ("Sample Description", 32), ("Device", 18), ("Lot Number", 16),
    ]
    for ci, (label, width) in enumerate(SUM_COLS, start=1):
        _hdr(ws_sum, 1, ci, label, width)
    ws_sum.row_dimensions[1].height = 22

    for ri, req in enumerate(requests, start=2):
        vals = [
            req.get("ca_number"), req.get("title"), req.get("status"),
            req.get("submitter_name"), req.get("analyst_name"), req.get("priority"),
            req.get("created_at"), req.get("completed_at"),
            req.get("sample_description"), req.get("device"), req.get("lot_number"),
        ]
        for ci, v in enumerate(vals, start=1):
            _cell(ws_sum, ri, ci, v)

    ws_sum.freeze_panes = "A2"

    # ── Sheets 2+: one per request ───────────────────────────────────────
    for req in requests:
        sheet_name = (req.get("ca_number") or str(req["id"]))[:31]
        ws = wb.create_sheet(title=sheet_name)

        # General info block
        INFO_FIELDS = [
            ("CA Number",         req.get("ca_number")),
            ("Title",             req.get("title")),
            ("Status",            req.get("status")),
            ("Priority",          req.get("priority")),
            ("Submitter",         req.get("submitter_name")),
            ("Analyst",           req.get("analyst_name")),
            ("Sample Description",req.get("sample_description")),
            ("Device",            req.get("device")),
            ("Lot Number",        req.get("lot_number")),
            ("Department",        req.get("department")),
            ("Note",              req.get("note")),
            ("Created At",        req.get("created_at")),
            ("Approved At",       req.get("approved_at")),
            ("Completed At",      req.get("completed_at")),
            ("Discontinued At",   req.get("discontinued_at")),
            ("Discontinue Reason",req.get("discontinue_reason")),
        ]
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 45
        row = 1
        title_c = ws.cell(row=row, column=1, value=f"CA Request — {req.get('ca_number','')}")
        title_c.font = Font(bold=True, size=13, color="4C1D95")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
        for label, val in INFO_FIELDS:
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = Font(bold=True, size=9); lc.fill = sub_fill
            lc.alignment = left_al; lc.border = thin_border
            vc = ws.cell(row=row, column=2, value=val)
            vc.alignment = left_al; vc.border = thin_border
            row += 1

        # Steps
        row += 1
        ws.cell(row=row, column=1, value="CA Steps").font = Font(bold=True, size=10, color="4C1D95")
        row += 1
        for sh, sw in [("Step", 28), ("Status", 14), ("Started At", 20), ("Completed At", 20)]:
            _hdr(ws, row, ["Step","Status","Started At","Completed At"].index(sh)+1, sh, sw)
        row += 1
        for step in req["steps"]:
            _cell(ws, row, 1, step.get("step_name"))
            _cell(ws, row, 2, step.get("status"))
            _cell(ws, row, 3, step.get("started_at"))
            _cell(ws, row, 4, step.get("completed_at"))
            row += 1

        checklist = req["checklist"]
        if not checklist:
            continue

        # Group checklist by leg
        legs_order = []
        legs_map = {}
        for item in checklist:
            ln = item.get("leg_name") or ""
            lt = item.get("leg_title") or ""
            key = ln
            if key not in legs_map:
                legs_order.append(key)
                legs_map[key] = {"title": lt, "items": []}
            legs_map[key]["items"].append(item)

        # Checklist header
        row += 1
        ws.cell(row=row, column=1, value="Checklist").font = Font(bold=True, size=10, color="4C1D95")
        row += 1

        CL_COLS = ["Step", "Item", "Requirements", "Qty", "Time In", "Time Out", "Technician", "Remarks"]
        CL_WIDTHS = [22, 28, 36, 8, 18, 18, 18, 32]
        for ci, (sh, sw) in enumerate(zip(CL_COLS, CL_WIDTHS), start=1):
            ws.column_dimensions[ws.cell(row=row, column=ci).column_letter].width = sw
            _hdr(ws, row, ci, sh)
        row += 1

        for leg_key in legs_order:
            leg_info = legs_map[leg_key]
            leg_label = leg_key or "(No Leg)"
            if leg_info["title"]:
                leg_label += f" — {leg_info['title']}"
            # Leg header row
            for ci in range(1, len(CL_COLS)+1):
                c = ws.cell(row=row, column=ci)
                if ci == 1:
                    c.value = leg_label
                c.font = leg_font; c.fill = leg_fill
                c.alignment = left_al; c.border = thin_border
            row += 1
            for it in leg_info["items"]:
                _cell(ws, row, 1, it.get("step_name"))
                _cell(ws, row, 2, it.get("item_name"))
                _cell(ws, row, 3, it.get("requirements"))
                _cell(ws, row, 4, it.get("qty"))
                _cell(ws, row, 5, it.get("time_in"))
                _cell(ws, row, 6, it.get("time_out"))
                _cell(ws, row, 7, it.get("technician"))
                _cell(ws, row, 8, it.get("remarks"))
                row += 1

    # Build the file in memory first — delete from DB only after a clean save
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    exported_count = len(requests)

    # Permanently remove the exported completed requests from the database
    if requests:
        req_ids = [req["id"] for req in requests]
        placeholders = ",".join("?" * len(req_ids))
        async with aiosqlite.connect(DB_PATH) as db_del:
            # Explicit child-table deletes (safe even if CASCADE is not enforced)
            await db_del.execute(f"DELETE FROM ca_checklist_items WHERE request_id IN ({placeholders})", req_ids)
            await db_del.execute(f"DELETE FROM ca_steps         WHERE request_id IN ({placeholders})", req_ids)
            await db_del.execute(f"DELETE FROM ca_schedule      WHERE request_id IN ({placeholders})", req_ids)
            await db_del.execute(f"DELETE FROM ca_requests      WHERE id          IN ({placeholders})", req_ids)
            await db_del.commit()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"CA_Backup_{ts}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Exported-Count":    str(exported_count),
        },
    )


@app.post("/api/backup/import")
async def import_backup(file: UploadFile = File(...), _=Depends(require_role("Admin"))):
    """Read a CA backup Excel file and return its contents for preview — does NOT
    modify the database.  The first sheet (Summary) drives the request list;
    all detail sheets are parsed and returned for client-side display."""
    if not _OPENPYXL_OK:
        raise HTTPException(status_code=500, detail="openpyxl not installed on server")

    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Must be an Excel file (.xlsx)")

    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    # Parse Summary sheet
    ws_sum = wb.worksheets[0] if wb.worksheets else None
    requests_out = []
    if ws_sum:
        rows = list(ws_sum.iter_rows(values_only=True))
        if rows:
            headers = [str(h).strip() if h else "" for h in rows[0]]
            for row in rows[1:]:
                if not any(c for c in row):
                    continue
                d = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row) if i < len(headers)}
                if d.get("CA Number"):
                    requests_out.append({
                        "ca_number":          d.get("CA Number", ""),
                        "title":              d.get("Title", ""),
                        "status":             d.get("Status", ""),
                        "submitter_name":     d.get("Submitter", ""),
                        "analyst_name":       d.get("Analyst", ""),
                        "priority":           d.get("Priority", ""),
                        "created_at":         d.get("Created At", ""),
                        "completed_at":       d.get("Completed At", ""),
                        "sample_description": d.get("Sample Description", ""),
                        "device":             d.get("Device", ""),
                        "lot_number":         d.get("Lot Number", ""),
                        "detail_sheet":       None,
                    })

    # Parse detail sheets — map ca_number -> detail dict
    detail_map = {}
    for ws in wb.worksheets[1:]:
        ca_num = ws.title.strip()
        rows_all = list(ws.iter_rows(values_only=True))
        info = {}
        steps = []
        checklist_legs = {}   # leg_name -> list of items
        checklist_order = []
        mode = "info"  # "info" | "steps" | "checklist"
        step_hdr_seen = False
        cl_hdr_seen   = False
        cl_hdr_cols   = []
        step_hdr_cols = []
        for row in rows_all:
            vals = [str(c).strip() if c is not None else "" for c in row]
            non_empty = [v for v in vals if v]
            if not non_empty:
                continue
            joined = " ".join(non_empty).lower()
            # detect section switches
            if non_empty[0].lower().startswith("ca steps"):
                mode = "steps"; step_hdr_seen = False; continue
            if non_empty[0].lower().startswith("checklist"):
                mode = "checklist"; cl_hdr_seen = False; continue
            if mode == "info":
                if len(row) >= 2 and vals[0] and not vals[0].lower().startswith("ca request"):
                    info[vals[0]] = vals[1] if len(vals) > 1 else ""
            elif mode == "steps":
                if not step_hdr_seen:
                    step_hdr_cols = vals; step_hdr_seen = True; continue
                step_d = {step_hdr_cols[i]: vals[i] for i in range(min(len(step_hdr_cols), len(vals)))}
                if step_d.get("Step"):
                    steps.append(step_d)
            elif mode == "checklist":
                if not cl_hdr_seen:
                    if "Item" in vals or "item" in joined:
                        cl_hdr_cols = vals; cl_hdr_seen = True
                    continue
                # Check if it's a leg header (all cells same fill — we detect by content pattern)
                first_val = vals[0] if vals else ""
                if all(v == "" for v in vals[1:]):
                    # Single-column row = leg label
                    leg_key = first_val
                    if leg_key not in checklist_legs:
                        checklist_order.append(leg_key)
                        checklist_legs[leg_key] = []
                    continue
                # Regular checklist row
                if cl_hdr_cols:
                    it = {cl_hdr_cols[i]: vals[i] for i in range(min(len(cl_hdr_cols), len(vals)))}
                    if it.get("Item") or it.get("Step"):
                        leg_key = checklist_order[-1] if checklist_order else ""
                        if leg_key not in checklist_legs:
                            checklist_order.append(leg_key)
                            checklist_legs[leg_key] = []
                        checklist_legs[leg_key].append(it)
        detail_map[ca_num] = {"info": info, "steps": steps,
                               "checklist_legs": checklist_legs,
                               "checklist_order": checklist_order}

    # Attach detail to each request
    for req in requests_out:
        cn = req["ca_number"]
        req["detail"] = detail_map.get(cn)

    return {
        "requests": requests_out,
        "sheets": list(detail_map.keys()),
        "file_name": file.filename,
    }


# ── Settings ───────────────────────────────────────────────────────────────
@app.get("/api/settings")
async def get_settings(_=Depends(get_current_user)):
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        cur = await db.execute("SELECT key,value FROM settings")
        rows = await cur.fetchall()
    return {r["key"]: r["value"] for r in rows}

@app.put("/api/settings")
async def put_settings(body: dict, _=Depends(require_role("Admin"))):
    async with aiosqlite.connect(DB_PATH) as db:
        for k, v in body.items():
            await db.execute("INSERT OR REPLACE INTO settings (key,value) VALUES (?,?)", (k, str(v)))
        await db.commit()
    return {"message": "Settings saved"}

# ── Serve frontend static files if built (production / LAN mode) ──────────
frontend_dist = ROOT_DIR.parent / "frontend" / "dist"
if frontend_dist.exists():
    _assets = frontend_dist / "assets"
    if _assets.exists():
        app.mount("/assets", StaticFiles(directory=str(_assets)), name="ca-frontend-assets")

# Always register catch-all so browsers never see raw FastAPI 404
@app.get("/{full_path:path}", include_in_schema=False)
async def serve_spa(full_path: str):
    """Catch-all: serve built SPA or helpful hint when dist is missing."""
    if frontend_dist.exists():
        file_candidate = frontend_dist / full_path
        if file_candidate.is_file():
            return FileResponse(str(file_candidate))
        return FileResponse(str(frontend_dist / "index.html"))
    # Frontend not built yet – give a useful message
    from fastapi.responses import JSONResponse
    return JSONResponse(
        status_code=200,
        content={
            "message": "CA Request API is running. Frontend not built yet.",
            "hint": "Run 'npm run build' in ca-website/frontend, or use Vite dev server on port 3001.",
            "api_health": "/api/health"
        }
    )

# ── Run ────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    # ── Load configuration from environment
    workers = int(os.getenv("WORKERS", "1"))
    reload = os.getenv("RELOAD", "false").lower() == "true"
    
    logger.info(f"CA Server starting on {HOST}:{PORT}...")
    uvicorn.run(
        "server:app",
        host=HOST,
        port=PORT,
        workers=workers,
        reload=reload,
        log_level="info"
    )

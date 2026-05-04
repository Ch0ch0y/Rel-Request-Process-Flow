"""Generate LTC directly using the server's _generate_ltc_excel function and check F8."""
import sys
import sqlite3
import json
from pathlib import Path
from datetime import datetime, timezone
import io
import zipfile
import re

# Import directly from server
sys.path.insert(0, str(Path('Rel Website/backend')))

# Get request data from DB
conn = sqlite3.connect('Rel Website/backend/rel_database.db')
conn.row_factory = sqlite3.Row
cur = conn.execute("SELECT * FROM requests WHERE request_number = ?", ("RR00143406",))
row = cur.fetchone()
if not row:
    print("Request not found")
    sys.exit(1)

# Build req dict
cols = [desc[0] for desc in cur.description]
req = dict(zip(cols, row))
req['automotive'] = bool(req.get('automotive'))

# Get steps
step_cur = conn.execute(
    "SELECT * FROM process_steps WHERE request_id = ? ORDER BY leg, step_number",
    (req['id'],)
)
step_cols = [desc[0] for desc in step_cur.description]
steps = []
for s in step_cur.fetchall():
    sd = dict(zip(step_cols, s))
    sd['custom_fields'] = json.loads(sd['custom_fields']) if sd.get('custom_fields') else {}
    sd['attachments'] = json.loads(sd['attachments']) if sd.get('attachments') else []
    steps.append(sd)
req['steps'] = steps
conn.close()

print(f"Request: request_number={req['request_number']!r}")
print(f"         original_rr_number={req.get('original_rr_number')!r}")
print(f"         Steps: {len(steps)}")

# Now call _generate_ltc_excel
from openpyxl import load_workbook

TPL = Path('Rel Website/backend/templates/REL LTC Template.xlsx')
wb = load_workbook(str(TPL))
ws = wb['Sheet1']

from openpyxl.cell.cell import MergedCell

# Replicate the fixed F8 logic
req_num = req.get("request_number", "") or ""
auto_rel = req.get("original_rr_number") or req_num
print(f"\nauto_rel = {auto_rel!r}")

# Write F8 directly (as the fixed code does)
cell_f8 = ws["F8"]
print(f"F8 before write: value={cell_f8.value!r}, is_merged={isinstance(cell_f8, MergedCell)}")
cell_f8.value = auto_rel or ""
print(f"F8 after write:  value={cell_f8.value!r}")

# Save and do ZIP rebuild exactly like server
buf = io.BytesIO()
wb.save(buf)
buf.seek(0)

with zipfile.ZipFile(buf) as gen_zip:
    gen_names = gen_zip.namelist()
    sheet_xml = gen_zip.read('xl/worksheets/sheet1.xml')
    shared_strings = gen_zip.read('xl/sharedStrings.xml') if 'xl/sharedStrings.xml' in gen_names else None
    styles_xml = gen_zip.read('xl/styles.xml') if 'xl/styles.xml' in gen_names else None

print(f"\nshared_strings present: {shared_strings is not None}")
print(f"styles_xml present: {styles_xml is not None}")

# Check F8 in the sheet XML
f8_match = re.search(rb'<c r="F8"[^/]*(?:/>|>.*?</c>)', sheet_xml, re.DOTALL)
print(f"F8 in openpyxl sheet XML: {f8_match.group(0).decode() if f8_match else 'NOT FOUND'}")

# Rebuild ZIP
output = io.BytesIO()
with zipfile.ZipFile(str(TPL), 'r') as tpl_zip, \
     zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as out_zip:
    for name in tpl_zip.namelist():
        if name == 'xl/worksheets/sheet1.xml':
            out_zip.writestr(name, sheet_xml)
        elif name == 'xl/sharedStrings.xml' and shared_strings is not None:
            out_zip.writestr(name, shared_strings)
        elif name == 'xl/styles.xml' and styles_xml is not None:
            out_zip.writestr(name, styles_xml)
        elif name == 'xl/calcChain.xml':
            pass
        elif name == '[Content_Types].xml':
            ct = tpl_zip.read(name)
            ct = re.sub(rb'<Override\s+PartName="/xl/calcChain\.xml"[^>]*/>', b'', ct)
            out_zip.writestr(name, ct)
        elif name == 'xl/_rels/workbook.xml.rels':
            rels = tpl_zip.read(name)
            rels = re.sub(rb'<Relationship\s[^>]*calcChain[^>]*/>', b'', rels)
            out_zip.writestr(name, rels)
        else:
            out_zip.writestr(name, tpl_zip.read(name))

output.seek(0)
final_data = output.getvalue()

# Check final output
with zipfile.ZipFile(io.BytesIO(final_data)) as z:
    final_sheet = z.read('xl/worksheets/sheet1.xml').decode()
    f8_final = re.search(r'<c r="F8"[^/]*(?:/>|>.*?</c>)', final_sheet, re.DOTALL)
    d8_final = re.search(r'<c r="D8"[^/]*(?:/>|>.*?</c>)', final_sheet, re.DOTALL)
    print(f"\nFINAL D8 XML: {d8_final.group(0) if d8_final else 'NOT FOUND'}")
    print(f"FINAL F8 XML: {f8_final.group(0) if f8_final else 'NOT FOUND'}")
    
    # Check if shared strings still has the old value
    if 'xl/sharedStrings.xml' in z.namelist():
        ss = z.read('xl/sharedStrings.xml').decode()
        if 'RRS# 220260256' in ss:
            print("\nWARNING: Old template sharedStrings with 'RRS# 220260256' is in the final ZIP")
        if 'REL202600001' in ss:
            print("OK: REL202600001 in sharedStrings")
    else:
        print("\nNo sharedStrings.xml in final ZIP")

# Save for inspection
Path('test_ltc_f8_check.xlsx').write_bytes(final_data)
print(f"\nSaved to test_ltc_f8_check.xlsx ({len(final_data)} bytes)")
print("Open it in Excel and check F8!")

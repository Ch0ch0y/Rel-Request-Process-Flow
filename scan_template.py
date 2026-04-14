"""Scan REL LTC Template using openpyxl to map all non-empty cells."""
from openpyxl import load_workbook

wb = load_workbook('Rel Website/backend/templates/REL LTC Template.xlsx', data_only=True)
ws = wb['Sheet1']

print("=== NON-EMPTY CELLS (rows 1-60) ===")
for row in ws.iter_rows(min_row=1, max_row=60):
    for cell in row:
        if cell.value is not None and str(cell.value).strip() != '':
            print(f"  {cell.coordinate:6s}: {repr(str(cell.value))[:80]}")

print()
print("=== MERGED CELL RANGES ===")
for m in ws.merged_cells.ranges:
    print(f"  {m}")

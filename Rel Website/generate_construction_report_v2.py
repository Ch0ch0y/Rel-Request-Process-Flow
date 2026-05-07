import datetime
import openpyxl
import os
from openpyxl.styles import Alignment, Border, Side, Font

# Configuration
START_DATE = datetime.date(2026, 2, 12)
END_DATE = datetime.date(2026, 5, 6)
TEMPLATE_PATH = r"c:\Users\168056\OneDrive - Amkor Technology\Documents\GitHub\Rel Request Process Flow\Rel Website\Apprentice_Weekly_Report_RELDMS.xlsx"
OUTPUT_PATH = r"c:\Users\168056\OneDrive - Amkor Technology\Documents\GitHub\Rel Request Process Flow\Rel Website\Construction_Report_Final.xlsx"

# Philippine Holidays 2026 (Feb 12 to May 6)
HOLIDAYS = {
    datetime.date(2026, 2, 17),  # Chinese New Year
    datetime.date(2026, 2, 25),  # EDSA People Power Revolution Anniversary
    datetime.date(2026, 3, 20),  # Eid al-Fitr (Tentative)
    datetime.date(2026, 4, 2),   # Maundy Thursday
    datetime.date(2026, 4, 3),   # Good Friday
    datetime.date(2026, 4, 4),   # Black Saturday (Included in prompt skip)
    datetime.date(2026, 4, 9),   # Araw ng Kagitingan
    datetime.date(2026, 5, 1),   # Labor Day
}

COMMITS = [
    ("2026-05-04", "Add files via upload - Finalizing assets and components"),
    ("2026-04-30", "LTC footer implementation, 24hr timing logic, and launcher stability fixes"),
    ("2026-04-24", "Backend and frontend synchronization and state management update"),
    ("2026-04-20", "Project documentation and system architectural overview"),
    ("2026-04-17", "Database schema fixes and RMS dropdown visibility improvements"),
    ("2026-04-15", "Agile import duplicate resolution and Render deployment scripts"),
    ("2026-04-14", "Agile RSS import enhancements (multi-sheet matrix scan)"),
    ("2026-04-10", "Process Monitoring page and Masterlist enhancements"),
    ("2026-04-06", "RelMon integration: custom device sheets (ATP1/ATP3)"),
    ("2026-04-01", "Test Level autocomplete and View All Device Type feature"),
    ("2026-03-31", "Production readiness: Render deployment, responsive UI, launcher cleanup"),
]

def is_workday(d):
    if d.weekday() >= 5: # 5=Saturday, 6=Sunday
        return False
    if d in HOLIDAYS:
        return False
    return True

def get_activity_data(d):
    ds = d.strftime("%Y-%m-%d")
    
    # Check for exact commit
    for date_str, msg in COMMITS:
        if date_str == ds:
            return {
                "title": "RELDMS Construction: " + msg,
                "background": "Developing internal REL Management system to digitize request flow.",
                "objective": "Implement specific feature sets or bug fixes identified in the roadmap.",
                "status": "Completed / Verified",
                "learnings": f"Learned how to implement {msg.split(' - ')[0]} and handle technical constraints."
            }
    
    # General activities for days without specific commits
    if d < datetime.date(2026, 3, 31):
        titles = [
            "Backend API development and Database schema design",
            "Frontend React component scaffolding and Tailwind styling",
            "Integration of JWT Authentication and user roles",
            "UI/UX refinement for the Request Submission forms"
        ]
        idx = d.day % len(titles)
        return {
            "title": f"RELDMS Construction: {titles[idx]}",
            "background": "Initial build phase focused on core architecture and CRUD operations.",
            "objective": "Build a robust foundation for the Reliability Engineering request tracker.",
            "status": "In Progress",
            "learnings": "Gained experience in Full-stack coordination and FastAPI/React integration."
        }
    else:
        titles = [
            "System stabilization and edge-case bug fixing",
            "Performance monitoring and database query optimization",
            "Cross-browser compatibility and mobile responsiveness testing",
            "User acceptance testing (UAT) and feedback implementation"
        ]
        idx = d.day % len(titles)
        return {
            "title": f"RELDMS Construction: {titles[idx]}",
            "background": "Final touch and optimization phase before release.",
            "objective": "Ensure the system is high-performing and user-friendly.",
            "status": "In Progress / Optimization",
            "learnings": "Focused on deployment strategies and production-level system hardening."
        }

def fill_report():
    if not os.path.exists(TEMPLATE_PATH):
        print(f"Error: Template not found at {TEMPLATE_PATH}")
        return

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    def safe_write(r, c, value):
        # find the master cell of a merged range
        target_cell = ws.cell(row=r, column=c)
        for merged_range in ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
        target_cell.value = value

    # Define labels we are looking for in Column C
    LABELS = {
        "Activity No./ Title": "title",
        "Background": "background",
        "Objective": "objective",
        "Status": "status",
        "Learnings": "learnings"
    }

    # First, collect all workdays
    dates = []
    d = START_DATE
    while d <= END_DATE:
        if is_workday(d):
            dates.append(d)
        d += datetime.timedelta(days=1)

    date_index = 0
    # Scan Column C for labels and fill Column E
    # We scan a large enough range to cover multiple pages
    max_row = ws.max_row if ws.max_row > 500 else 1000
    
    current_date_row = -1
    
    for r in range(1, max_row + 1):
        if date_index >= len(dates):
            break
            
        # Get value from Column C (Master cell might be C or something else if merged)
        # But based on scan, labels were appearing in Column C
        cell_val = ws.cell(row=r, column=3).value
        if not cell_val:
            continue
            
        cell_text = str(cell_val).strip()
        
        # Check if this row matches one of our labels
        found_key = None
        for label_text, key in LABELS.items():
            if label_text in cell_text:
                found_key = key
                break
        
        if found_key:
            # We found a label row. Now get the data for the current date.
            data = get_activity_data(dates[date_index])
            
            # Place the Label (from LABELS dictionary key) into Column C
            # Using the original label_text found during scan
            label_display = next(k for k, v in LABELS.items() if v == found_key)
            safe_write(r, 3, label_display)
            
            # Write the Data info to Column E (index 5)
            safe_write(r, 5, data[found_key])
            
            # If we just wrote the "Learnings" (last item), move to the next date
            if found_key == "learnings":
                date_index += 1
                current_date_row = -1 # Reset for next block
        
        # Also handle the Date (Day) column if possible
        # Look for the date/Day label or typical position
        # In the scan, Column A (index 1) contained the date info
        day_val = ws.cell(row=r, column=1).value
        # If it's a date-like row and we haven't set the date for this block yet
        if found_key == "title" and date_index < len(dates):
             # Ensure the date is written in Column A of the title row
             safe_write(r, 1, dates[date_index].strftime("%m/%d/%y"))

    wb.save(OUTPUT_PATH)
    print(f"Generated report at {OUTPUT_PATH}")

if __name__ == "__main__":
    fill_report()

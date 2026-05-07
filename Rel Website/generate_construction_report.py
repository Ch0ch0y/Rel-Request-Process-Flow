import datetime
import openpyxl
from openpyxl.styles import Alignment, Border, Side

# Configuration
START_DATE = datetime.date(2026, 2, 12)
END_DATE = datetime.date(2026, 5, 6)
TEMPLATE_PATH = r"c:\Users\168056\OneDrive - Amkor Technology\Documents\GitHub\Rel Request Process Flow\Rel Website\Sample Docs\Apprentice Weekly Report.xlsx"
OUTPUT_PATH = r"c:\Users\168056\OneDrive - Amkor Technology\Documents\GitHub\Rel Request Process Flow\Rel Website\Apprentice_Weekly_Report_RELDMS.xlsx"

HOLIDAYS = {
    datetime.date(2026, 2, 17), # Chinese New Year
    datetime.date(2026, 2, 25), # EDSA
    datetime.date(2026, 3, 20), # Eid al-Fitr
    datetime.date(2026, 4, 2),  # Maundy Thursday
    datetime.date(2026, 4, 3),  # Good Friday
    datetime.date(2026, 4, 4),  # Black Saturday
    datetime.date(2026, 5, 1),  # Labor Day
}

def is_workday(d):
    if d.weekday() >= 5: # Saturday=5, Sunday=6
        return False
    if d in HOLIDAYS:
        return False
    return True

# Map commits to dates
commits = [
    ("2026-05-06", "Add SAT Sonoscan workbench and activity tracking"),
    ("2026-05-04", "Frontend UI refinements and file upload fixes"),
    ("2026-04-30", "LTC footer implementation, 24hr timing logic, launcher fixes"),
    ("2026-04-24", "Backend and frontend synchronization"),
    ("2026-04-20", "Project documentation and README updates"),
    ("2026-04-17", "Database schema fixes and RMS dropdown visibility improvements"),
    ("2026-04-15", "Agile import duplicate resolution and Render deployment scripts"),
    ("2026-04-14", "Agile RSS import enhancements (multi-sheet matrix scan)"),
    ("2026-04-10", "Process Monitoring page and Masterlist enhancements"),
    ("2026-04-06", "RelMon integration: custom device sheets (ATP1/ATP3)"),
    ("2026-04-01", "Test Level autocomplete and View All Device Type feature"),
    ("2026-03-31", "Production readiness: Render deployment, responsive UI, launcher cleanup"),
    ("2026-02-11", "Initial project setup and repository creation"),
]

def get_activity(d):
    ds = d.strftime("%Y-%m-%d")
    for date_str, msg in commits:
        if date_str == ds:
            return msg
    if d < datetime.date(2026, 3, 31):
        if d.day % 3 == 0: return "Backend API development and database optimization"
        if d.day % 3 == 1: return "Frontend component construction and styling"
        return "Integration testing and bug fixing"
    elif d < datetime.date(2026, 5, 6):
        if d.day % 2 == 0: return "Feature enhancement and UI polishing"
        return "System stabilization and deployment testing"
    return "RELDMS Construction"

def main():
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Helper function to write to merged cells (write to the top-left cell)
    def write_cell_rc(r, c, value):
        from openpyxl.utils.cell import get_column_letter
        coord = f"{get_column_letter(c)}{r}"
        
        # Check if the cell is part of a merged range
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                # Write to the top-left cell of the merged range
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
        
        # If not merged, write directly
        ws.cell(row=r, column=c).value = value

    def write_cell(coord, value):
        from openpyxl.utils.cell import coordinate_to_tuple
        r, c = coordinate_to_tuple(coord)
        write_cell_rc(r, c, value)

    # Personal Info (Based on template structure)
    # The template has very large merged ranges (e.g., A5:Q5 contains Name, ID, etc labels)
    # We must write to the specific cells that are NOT part of labels if possible, 
    # or identify the exact master cells.
    
    # According to the image:
    # B5 prefix is "Name: ", we want to write to the space after it.
    # However, since A5:Q5 is merged, writing to any part of it hits Column A.
    
    # Let's try to write the header info more carefully.
    write_cell_rc(5, 3, "Francis Niño Villanueva") # Attempt to write in the Name area
    write_cell_rc(6, 3, "631090")                  # ID
    write_cell_rc(7, 3, "Dec 17, 2025")           # Date
    
    write_cell_rc(5, 6, "Apprentice Engineer")    # Position
    write_cell_rc(6, 6, "RE_FA")                  # Station
    write_cell_rc(7, 6, "RELLAB")                 # Dept
    
    write_cell_rc(5, 10, f"{START_DATE.strftime('%m/%d/%y')} - {END_DATE.strftime('%m/%d/%y')}")
    write_cell_rc(6, 10, "Loreta Veran")
    write_cell_rc(7, 10, "Rodam Lopez")

    def find_row_for_label(start_row, label_text):
        """Find the row index within a range that contains the label in Column C."""
        for r in range(start_row, start_row + 50): # Scan up to 50 rows
            # Based on the latest inspection, labels are in Column C
            cell_c = ws.cell(row=r, column=3).value
            if cell_c and label_text in str(cell_c):
                return r
        return None

    curr_row = 1 
    curr_date = START_DATE
    
    while curr_date <= END_DATE:
        if is_workday(curr_date):
            # Scan for 'Activity No./ Title' starting from curr_row
            target_row = find_row_for_label(curr_row, "Activity No./ Title")
            if target_row is None:
                break 
            
            activity = get_activity(curr_date)
            bg = "Continuous development of RELDMS platform."
            obj = "Enhance system functionality and user experience."
            status = "Completed"
            learn = f"Improved understanding of {'React/Vite' if curr_date.day % 2 == 0 else 'FastAPI/SQLite'} integration."
            
            # Fill Column D (according to the blank highlight area)
            # We match the label rows found in the template
            row_act = find_row_for_label(target_row, "Activity No./ Title")
            row_bg = find_row_for_label(target_row, "Background")
            row_obj = find_row_for_label(target_row, "Objective")
            row_status = find_row_for_label(target_row, "Status")
            row_learn = find_row_for_label(target_row, "Learnings")

            # Writing to Column D (Index 4) - This is the "Highlights" data area
            if row_act: write_cell_rc(row_act, 4, activity)
            if row_bg: write_cell_rc(row_bg, 4, bg)
            if row_obj: write_cell_rc(row_obj, 4, obj)
            if row_status: write_cell_rc(row_status, 4, status)
            if row_learn: write_cell_rc(row_learn, 4, learn)

            # Date in Column A of the Activity row
            write_cell_rc(target_row, 1, curr_date.strftime('%m/%d/%Y'))

            # Stylize the data in Column D
            for r in [row_act, row_bg, row_obj, row_status, row_learn]:
                if r:
                    cell = ws.cell(row=r, column=4)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Move search pointer past this block (Learnings + space)
            if row_learn:
                curr_row = row_learn + 2
            else:
                curr_row = target_row + 10

        curr_date += datetime.timedelta(days=1)

        curr_date += datetime.timedelta(days=1)

    wb.save(OUTPUT_PATH)
    print(f"Report saved to: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()

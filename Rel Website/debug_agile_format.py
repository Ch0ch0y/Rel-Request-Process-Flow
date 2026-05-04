"""
debug_agile_format.py
─────────────────────
Run this against a real Agile RSS Report Excel file to show:
  1. The first 40 rows with ALL non-empty cells (row, col, value)
  2. A flat dump of what the current parser detects

Usage (from "Rel Website" folder):
  python debug_agile_format.py "path/to/your_agile_file.xlsx"
"""

import sys
import io
import openpyxl

def main():
    if len(sys.argv) < 2:
        print("Usage: python debug_agile_format.py <path_to_agile_excel.xlsx>")
        sys.exit(1)

    path = sys.argv[1]
    with open(path, 'rb') as f:
        contents = f.read()

    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    print(f"Sheets: {wb.sheetnames}\n")

    ws = wb.worksheets[0]
    print(f"Active sheet: {ws.title}  |  max_row={ws.max_row}  max_col={ws.max_column}\n")

    # ── 1. Dump first 50 rows, all non-empty cells ──────────────────────────
    print("=" * 70)
    print("NON-EMPTY CELLS (first 50 rows, all columns)")
    print("=" * 70)
    for r in range(1, min(51, ws.max_row + 1)):
        row_parts = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip():
                col_letter = openpyxl.utils.get_column_letter(c)
                row_parts.append(f"  {col_letter}{r}={repr(str(val).strip())}")
        if row_parts:
            print(f"Row {r:>3}:" + "".join(row_parts))

    # ── 2. Show what the parser currently detects ───────────────────────────
    print("\n" + "=" * 70)
    print("CURRENT PARSER OUTPUT (label->value using col B->D, col K->O)")
    print("=" * 70)

    def _cv(row_idx, col_idx):
        val = ws.cell(row=row_idx, column=col_idx).value
        return str(val).strip() if val is not None else ''

    general = {}
    for r in range(1, ws.max_row + 1):
        label       = _cv(r, 2).lower().strip()   # col B
        value       = _cv(r, 4)                   # col D
        right_label = _cv(r, 11).lower().strip()  # col K
        right_value = _cv(r, 15)                  # col O
        if label:
            general[label] = value
        if right_label:
            general[right_label] = right_value

    EXPECTED_KEYS = [
        'request number', 'classification', 'originator', 'plant',
        'device name', 'lot no', 'customer', 'pkg info', 'automotive',
        'purpose', 'total s/s', 'product hierarchy', 'pdl',
        'body size x (mm)', 'body size y (mm)', 'package thickness (mm)',
        'lead pitch (mm)', 'lead count', 'ball pitch (mm)', 'ball count',
    ]
    for key in EXPECTED_KEYS:
        val = general.get(key, '<NOT FOUND>')
        status = "OK" if val and val != '<NOT FOUND>' else "MISSING"
        print(f"  [{status:^7}]  {key:<30} = {repr(val)}")

    # ── 3. Show ALL detected general labels ────────────────────────────────
    print("\n" + "=" * 70)
    print("ALL LABELS DETECTED in col B (label->value from col D)")
    print("=" * 70)
    for r in range(1, ws.max_row + 1):
        label = _cv(r, 2).strip()
        value = _cv(r, 3)
        if label:
            print(f"  Row {r:>3}  B={repr(label):<40}  C={repr(value)}")

    print("\n" + "=" * 70)
    print("ALL LABELS DETECTED in col D (label->value from col E)")
    print("=" * 70)
    for r in range(1, ws.max_row + 1):
        label = _cv(r, 4).strip()
        value = _cv(r, 5)
        if label:
            print(f"  Row {r:>3}  D={repr(label):<40}  E={repr(value)}")


if __name__ == "__main__":
    main()
